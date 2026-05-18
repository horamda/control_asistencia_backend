"""
Exportador masivo de empleados en formato Excel compatible con el importador.

Genera el mismo layout que el template de importacion:
  Hoja 1 – Empleados: datos reales listos para editar y reimportar.
  Hoja 2 – Empresas, 3 – Sucursales, 4 – Sectores, 5 – Puestos,
  6 – Localidades, 7 – Valores validos  (mismas hojas de referencia).
"""

import io
import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extensions import get_db
from repositories.empresa_repository import get_all as get_empresas
from repositories.sucursal_repository import get_all as get_sucursales
from repositories.sector_repository import get_all as get_sectores
from repositories.puesto_repository import get_all as get_puestos
from repositories.localidad_repository import get_all as get_localidades


# ── Estilos (mismos que el template) ─────────────────────────────────────────

_HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT   = Font(bold=True, color="FFFFFF")
_REF_FILL      = PatternFill("solid", fgColor="2E75B6")
_REF_FONT      = Font(bold=True, color="FFFFFF")
_REQUIRED_FILL = PatternFill("solid", fgColor="FFE699")
_NOTE_FILL     = PatternFill("solid", fgColor="E2EFDA")
_NOTE_FONT     = Font(italic=True, color="375623")
_DATA_ALT_FILL = PatternFill("solid", fgColor="F2F9FF")


# ── Definicion de columnas (mismo orden que el importador) ────────────────────

# (clave_csv, label_legible, obligatorio, campo_db)
_COLUMNS = [
    ("legajo",            "Legajo",           True,  "legajo"),
    ("dni",               "DNI",              True,  "dni"),
    ("cuil",              "CUIL",             False, "cuil"),
    ("apellido",          "Apellido",         True,  "apellido"),
    ("nombre",            "Nombre",           True,  "nombre"),
    ("sexo",              "Sexo",             False, "sexo"),
    ("fecha_nacimiento",  "Fecha nacimiento", False, "fecha_nacimiento"),
    ("email",             "Email",            False, "email"),
    ("telefono",          "Telefono",         False, "telefono"),
    ("direccion",         "Direccion",        False, "direccion"),
    ("codigo_postal",     "Codigo postal",    False, "codigo_postal"),
    ("fecha_ingreso",     "Fecha ingreso",    False, "fecha_ingreso"),
    ("fecha_baja",        "Fecha baja",       False, "fecha_baja"),
    ("tipo_contrato",     "Tipo contrato",    False, "tipo_contrato"),
    ("modalidad",         "Modalidad",        False, "modalidad"),
    ("categoria",         "Categoria",        False, "categoria"),
    ("obra_social",       "Obra social",      False, "obra_social"),
    ("cod_chess_erp",     "Cod. Chess ERP",   False, "cod_chess_erp"),
    ("banco",             "Banco",            False, "banco"),
    ("cbu",               "CBU",              False, "cbu"),
    ("numero_emergencia", "Nro. emergencia",  False, "numero_emergencia"),
    ("estado",            "Estado",           False, "estado"),
    ("sucursal_nombre",   "Sucursal nombre",  False, "sucursal_nombre"),
    ("sector_nombre",     "Sector nombre",    False, "sector_nombre"),
    ("puesto_nombre",      "Puesto nombre",          False, "puesto_nombre"),
    ("reporta_a_legajo",  "Legajo jefe directo",    False, "reporta_a_legajo"),
    ("password",          "Password",               False, None),  # nunca se exporta
]

_DATE_FIELDS = {"fecha_nacimiento", "fecha_ingreso", "fecha_baja"}


# ── Query de empleados para exportar ──────────────────────────────────────────

def _get_empleados_for_export(
    empresa_id: int | None = None,
    activo: int | None = None,
) -> list[dict]:
    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        where = []
        params = []
        if empresa_id:
            where.append("e.empresa_id = %s")
            params.append(int(empresa_id))
        if activo in (0, 1):
            where.append("e.activo = %s")
            params.append(activo)
        where_sql = ("WHERE " + " AND ".join(where)) if where else ""
        cursor.execute(
            f"""
            SELECT
                e.legajo,
                e.dni,
                e.cuil,
                e.apellido,
                e.nombre,
                e.sexo,
                e.fecha_nacimiento,
                e.email,
                e.telefono,
                e.direccion,
                e.codigo_postal,
                e.fecha_ingreso,
                e.tipo_contrato,
                e.modalidad,
                e.fecha_baja,
                e.categoria,
                e.obra_social,
                e.cod_chess_erp,
                e.banco,
                e.cbu,
                e.numero_emergencia,
                e.estado,
                s.nombre   AS sucursal_nombre,
                sec.nombre AS sector_nombre,
                p.nombre   AS puesto_nombre,
                mgr.legajo AS reporta_a_legajo
            FROM empleados e
            LEFT JOIN sucursales s   ON s.id   = e.sucursal_id
            LEFT JOIN sectores   sec ON sec.id = e.sector_id
            LEFT JOIN puestos    p   ON p.id   = e.puesto_id
            LEFT JOIN empleados  mgr ON mgr.id = e.reporta_a_empleado_id
            {where_sql}
            ORDER BY e.apellido, e.nombre, e.id
            """,
            tuple(params),
        )
        return cursor.fetchall()
    finally:
        cursor.close()
        db.close()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fmt_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def _set_header_row(ws, cols: list[str], fill=None, font=None, row: int = 1):
    fill = fill or _HEADER_FILL
    font = font or _HEADER_FONT
    for col_idx, text in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws, min_width: int = 10, max_width: int = 45):
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value or "")) for c in col_cells),
            default=min_width,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(length + 3, min_width), max_width
        )


# ── Hoja principal: empleados ─────────────────────────────────────────────────

def _build_empleados(wb: Workbook, empleados: list[dict]):
    ws = wb.active
    ws.title = "Empleados"

    total = len(empleados)

    # Fila 1: instruccion
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))
    note = ws.cell(row=1, column=1)
    note.value = (
        f"EXPORTACION DE EMPLEADOS — {total} registro(s). "
        "Edite los datos directamente en esta hoja. "
        "Columnas en amarillo son OBLIGATORIAS para el importador. "
        "La columna Password se deja vacia; si la completa se usara como nueva contrasena al reimportar. "
        "Cuando termine, exporte ESTA HOJA como CSV (UTF-8) y carguela en el importador."
    )
    note.fill = _NOTE_FILL
    note.font = _NOTE_FONT
    note.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 40

    # Fila 2: claves CSV (para que el importador las reconozca)
    _set_header_row(ws, [c[0] for c in _COLUMNS], row=2)

    # Fila 3: etiquetas legibles
    for col_idx, (key, label, required, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=f"{label}{'  *' if required else ''}")
        if required:
            cell.fill = _REQUIRED_FILL
            cell.font = Font(bold=True, color="7B3F00")
        else:
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.font = Font(color="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # Filas de datos (desde fila 4)
    for row_idx, emp in enumerate(empleados, start=4):
        alt = (row_idx % 2 == 0)
        for col_idx, (csv_key, _, _, db_field) in enumerate(_COLUMNS, start=1):
            if db_field is None:
                value = ""  # password: nunca exportar
            elif db_field in _DATE_FIELDS:
                value = _fmt_date(emp.get(db_field))
            else:
                raw = emp.get(db_field)
                value = str(raw) if raw is not None else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if alt:
                cell.fill = _DATA_ALT_FILL

    ws.freeze_panes = "A4"
    _autofit(ws)


# ── Hojas de referencia (identicas al template) ───────────────────────────────

def _build_empresas(wb):
    ws = wb.create_sheet("Empresas")
    _set_header_row(ws, ["ID", "Razon social"], fill=_REF_FILL, font=_REF_FONT)
    for e in get_empresas(include_inactive=False):
        ws.append([e.get("id"), e.get("razon_social")])
    _autofit(ws)


def _build_sucursales(wb):
    ws = wb.create_sheet("Sucursales")
    _set_header_row(ws, ["Empresa", "Nombre (usar en CSV)", "ID"], fill=_REF_FILL, font=_REF_FONT)
    for s in get_sucursales(include_inactive=False):
        ws.append([s.get("empresa_nombre"), s.get("nombre"), s.get("id")])
    _autofit(ws)


def _build_sectores(wb):
    ws = wb.create_sheet("Sectores")
    _set_header_row(ws, ["Empresa", "Nombre (usar en CSV)", "ID"], fill=_REF_FILL, font=_REF_FONT)
    for s in get_sectores(include_inactive=False):
        ws.append([s.get("empresa_nombre"), s.get("nombre"), s.get("id")])
    _autofit(ws)


def _build_puestos(wb):
    ws = wb.create_sheet("Puestos")
    _set_header_row(ws, ["Empresa", "Nombre (usar en CSV)", "ID"], fill=_REF_FILL, font=_REF_FONT)
    for p in get_puestos(include_inactive=False):
        ws.append([p.get("empresa_nombre"), p.get("nombre"), p.get("id")])
    _autofit(ws)


def _build_localidades(wb):
    ws = wb.create_sheet("Localidades")
    _set_header_row(ws, ["Codigo postal (usar en CSV)", "Localidad", "Provincia"], fill=_REF_FILL, font=_REF_FONT)
    for loc in get_localidades():
        ws.append([loc.get("codigo_postal"), loc.get("localidad"), loc.get("provincia")])
    _autofit(ws)


def _build_valores_validos(wb):
    ws = wb.create_sheet("Valores validos")
    _set_header_row(ws, ["Campo CSV", "Valores aceptados", "Notas"], fill=_REF_FILL, font=_REF_FONT)
    rows = [
        ("sexo",           "masculino | femenino | no_binario | no_informa", "Default: no_informa"),
        ("estado",         "activo | inactivo | suspendido",                  "Default: activo"),
        ("tipo_contrato",  "efectivo | temporal | pasantia | otro",           "Opcional"),
        ("modalidad",      "presencial | remoto | hibrido",                   "Default: presencial"),
        ("password",       "(cualquier texto)",                                "Si se omite, se usa el DNI como contrasena inicial"),
        ("fecha_*",        "YYYY-MM-DD",                                      "Ejemplo: 2024-01-15"),
        ("cbu",              "22 digitos numericos",                          ""),
        ("cod_chess_erp",   "Numero entero",                                 ""),
        ("reporta_a_legajo","Legajo del empleado jefe directo",              "Debe existir en el sistema. Dejar vacio si no tiene jefe."),
    ]
    for r in rows:
        ws.append(list(r))
    _autofit(ws)


# ── Funcion publica ───────────────────────────────────────────────────────────

def exportar_empleados_excel(
    empresa_id: int | None = None,
    activo: int | None = None,
) -> bytes:
    """
    Genera el Excel de empleados con los datos actuales del sistema.
    empresa_id: filtra por empresa (None = todas).
    activo: 1=solo activos, 0=solo inactivos, None=todos.
    Devuelve bytes del .xlsx listo para servir.
    """
    empleados = _get_empleados_for_export(empresa_id=empresa_id, activo=activo)

    wb = Workbook()
    _build_empleados(wb, empleados)
    _build_empresas(wb)
    _build_sucursales(wb)
    _build_sectores(wb)
    _build_puestos(wb)
    _build_localidades(wb)
    _build_valores_validos(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
