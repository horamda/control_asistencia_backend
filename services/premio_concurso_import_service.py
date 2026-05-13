import csv
import datetime
import io
import unicodedata

from openpyxl import load_workbook

from repositories.premio_concurso_repository import (
    build_prepared_resultado,
    get_concursos_for_empresa,
    get_empleados_lookup_by_empresa,
    save_prepared_resultado,
)


_REQUIRED = {"periodo", "legajo", "codigo_concurso", "ranking"}
_ALIASES = {
    "periodo": "periodo",
    "mes": "periodo",
    "legajo": "legajo",
    "codigo_concurso": "codigo_concurso",
    "codigo concurso": "codigo_concurso",
    "codigo_premio": "codigo_concurso",
    "codigo premio": "codigo_concurso",
    "concurso": "codigo_concurso",
    "tipo_concurso": "codigo_concurso",
    "tipo concurso": "codigo_concurso",
    "tipo de concurso": "codigo_concurso",
    "ranking": "ranking",
    "posicion": "ranking",
    "puesto": "ranking",
    "observaciones": "observaciones",
    "observacion": "observaciones",
}


class PremioImportError(ValueError):
    pass


def _normalize_header(value) -> str:
    raw = str(value or "").strip().lower().lstrip("\ufeff")
    raw = "".join(
        char for char in unicodedata.normalize("NFKD", raw)
        if not unicodedata.combining(char)
    )
    raw = raw.replace("-", " ").replace("_", " ")
    raw = " ".join(raw.split())
    mapped = _ALIASES.get(raw)
    if mapped:
        return mapped
    return raw.replace(" ", "_")


def _normalize_codigo(value) -> str:
    return str(value or "").strip().upper()


def _parse_periodo(value) -> str:
    raw = str(value or "").strip()
    if not raw:
        raise ValueError("Periodo vacio")
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.replace(day=1).date().isoformat() if isinstance(value, datetime.datetime) else value.replace(day=1).isoformat()

    candidates = []
    if len(raw) == 7:
        candidates.append(f"{raw}-01")
    candidates.append(raw)
    for candidate in candidates:
        try:
            parsed = datetime.date.fromisoformat(candidate)
            return parsed.replace(day=1).isoformat()
        except ValueError:
            pass

    for fmt in ("%m/%Y", "%m-%Y"):
        try:
            parsed = datetime.datetime.strptime(raw, fmt).date()
            return parsed.replace(day=1).isoformat()
        except ValueError:
            pass
    raise ValueError(f"Periodo invalido: '{raw}'. Use YYYY-MM.")


def _parse_ranking(value) -> int:
    raw = str(value or "").strip()
    if not raw:
        raise ValueError("Ranking vacio")
    try:
        ranking = int(float(raw.replace(",", ".")))
    except ValueError as exc:
        raise ValueError(f"Ranking invalido: '{raw}'") from exc
    if ranking < 1:
        raise ValueError("Ranking debe ser mayor o igual a 1")
    return ranking


def _rows_from_csv(stream):
    raw = stream.read()
    text = raw.decode("utf-8-sig", errors="replace") if isinstance(raw, bytes) else str(raw or "")
    if not text.strip():
        raise PremioImportError("El archivo esta vacio.")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = {_normalize_header(h) for h in (reader.fieldnames or [])}
    missing = _REQUIRED - headers
    if missing:
        raise PremioImportError(f"Columnas faltantes: {', '.join(sorted(missing))}")
    for lineno, row in enumerate(reader, start=2):
        yield lineno, {_normalize_header(k): (v or "").strip() for k, v in row.items() if k is not None}


def _rows_from_xlsx(stream):
    workbook = load_workbook(stream, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise PremioImportError("El archivo esta vacio.") from exc
    headers = [_normalize_header(h) for h in raw_headers]
    header_set = {h for h in headers if h}
    missing = _REQUIRED - header_set
    if missing:
        raise PremioImportError(f"Columnas faltantes: {', '.join(sorted(missing))}")
    for lineno, values in enumerate(rows, start=2):
        row = {}
        for key, value in zip(headers, values):
            if not key:
                continue
            row[key] = value
        if not any(str(v or "").strip() for v in row.values()):
            continue
        yield lineno, row


def _iter_rows(stream, filename: str):
    lower = str(filename or "").lower()
    if lower.endswith(".xlsx"):
        yield from _rows_from_xlsx(stream)
        return
    if lower.endswith(".csv"):
        yield from _rows_from_csv(stream)
        return
    raise PremioImportError("Formato no soportado. Suba un archivo .csv o .xlsx.")


def importar_premios_desde_archivo(empresa_id: int, stream, filename: str, *, actor_id: int | None = None) -> dict:
    empleados_by_legajo = get_empleados_lookup_by_empresa(empresa_id)
    concursos_by_codigo = {
        _normalize_codigo(row.get("codigo")): row
        for row in get_concursos_for_empresa(empresa_id, activo=1)
    }

    total_filas = 0
    importadas = 0
    creados = 0
    actualizados = 0
    errors: list[dict] = []
    seen_empleado_periodo = set()
    seen_ranking = set()

    for lineno, row in _iter_rows(stream, filename):
        total_filas += 1
        legajo = str(row.get("legajo") or "").strip()
        codigo_concurso = _normalize_codigo(row.get("codigo_concurso"))
        observaciones = str(row.get("observaciones") or "").strip() or None

        try:
            periodo = _parse_periodo(row.get("periodo"))
            ranking = _parse_ranking(row.get("ranking"))
        except ValueError as exc:
            errors.append({"linea": lineno, "error": str(exc)})
            continue

        empleado = empleados_by_legajo.get(legajo)
        if not empleado:
            errors.append({"linea": lineno, "error": f"Legajo no encontrado: '{legajo}'"})
            continue

        concurso = concursos_by_codigo.get(codigo_concurso)
        if not concurso:
            errors.append({"linea": lineno, "error": f"Concurso no encontrado o inactivo: '{codigo_concurso}'"})
            continue

        empleado_key = (int(empleado["id"]), int(concurso["id"]), periodo)
        ranking_key = (int(concurso["id"]), periodo, ranking)
        if empleado_key in seen_empleado_periodo:
            errors.append({"linea": lineno, "error": "Duplicado en archivo para empleado, concurso y periodo."})
            continue
        if ranking_key in seen_ranking:
            errors.append({"linea": lineno, "error": "Duplicado en archivo para concurso, periodo y ranking."})
            continue

        try:
            prepared = build_prepared_resultado(
                empresa_id=empresa_id,
                empleado=empleado,
                concurso=concurso,
                periodo=periodo,
                ranking=ranking,
                observaciones=observaciones,
                actor_id=actor_id,
            )
            _, created = save_prepared_resultado(prepared)
        except ValueError as exc:
            errors.append({"linea": lineno, "error": str(exc)})
            continue

        seen_empleado_periodo.add(empleado_key)
        seen_ranking.add(ranking_key)
        importadas += 1
        if created:
            creados += 1
        else:
            actualizados += 1

    return {
        "total_filas": total_filas,
        "importadas": importadas,
        "creados": creados,
        "actualizados": actualizados,
        "errores": len(errors),
        "detalle_errores": errors[:100],
    }
