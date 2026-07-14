from __future__ import annotations

import datetime
import json
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Any, Iterable, Mapping, Sequence

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PDF_MIME = "application/pdf"

_TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
_TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
_HEADER_FILL = PatternFill("solid", fgColor="2F5597")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
_EMPTY_FONT = Font(italic=True, color="6B7280")

_KNOWN_ACRONYMS = {
    "api",
    "app",
    "csv",
    "cuit",
    "db",
    "dni",
    "gps",
    "id",
    "ip",
    "kpi",
    "pdf",
    "qr",
    "rrhh",
    "sla",
    "url",
    "web",
    "xls",
    "xlsx",
}


@dataclass(frozen=True)
class ExportFile:
    data: bytes
    mimetype: str
    filename: str


def normalize_export_format(raw: str | None) -> str:
    value = str(raw or "").strip().lower()
    if value in {"xls", "excel"}:
        return "xlsx"
    return value


def _pretty_label(raw_key: str) -> str:
    parts = [part for part in str(raw_key or "").replace("-", "_").split("_") if part]
    if not parts:
        return str(raw_key or "").strip() or "Dato"
    formatted = []
    for part in parts:
        lower = part.lower()
        if lower in _KNOWN_ACRONYMS:
            formatted.append(lower.upper())
        else:
            formatted.append(lower.capitalize())
    return " ".join(formatted)


def _normalize_excel_value(value: Any):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "SI" if value else "NO"
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value
    if isinstance(value, datetime.timedelta):
        total_seconds = int(value.total_seconds())
        sign = "-" if total_seconds < 0 else ""
        total_seconds = abs(total_seconds)
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        if seconds:
            return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"
        return f"{sign}{hours:02d}:{minutes:02d}"
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _normalize_pdf_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "SI" if value else "NO"
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, datetime.timedelta):
        total_seconds = int(value.total_seconds())
        sign = "-" if total_seconds < 0 else ""
        total_seconds = abs(total_seconds)
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        if seconds:
            return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"
        return f"{sign}{hours:02d}:{minutes:02d}"
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _coerce_row_mapping(row: Any) -> dict[str, Any]:
    if isinstance(row, Mapping):
        return dict(row)
    if isinstance(row, Sequence) and not isinstance(row, (str, bytes, bytearray)):
        return {str(index): value for index, value in enumerate(row, start=1)}
    return {"valor": row}


def _collect_columns(rows: Iterable[Any]) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        mapping = _coerce_row_mapping(row)
        for key in mapping.keys():
            key_str = str(key)
            if key_str not in seen:
                seen.add(key_str)
                columns.append(key_str)
    return columns


def _sheet_name(title: str) -> str:
    cleaned = "".join(ch for ch in (title or "Exportacion") if ch.isalnum() or ch in {" ", "_", "-"}).strip()
    cleaned = cleaned.replace("-", " ").replace("  ", " ")
    if not cleaned:
        cleaned = "Exportacion"
    return cleaned[:31]


def _autofit_columns(ws, *, max_width: int = 44):
    for column_cells in ws.columns:
        values = [str(cell.value or "") for cell in column_cells]
        if not values:
            continue
        width = max(len(value) for value in values)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 10), max_width)


def _build_xlsx(title: str, rows: list[dict[str, Any]]) -> bytes:
    columns = _collect_columns(rows)
    if not columns:
        columns = ["valor"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_name(title)

    end_col = max(1, len(columns))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    title_cell = sheet.cell(row=1, column=1, value=title or "Exportacion")
    title_cell.fill = _TITLE_FILL
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_row = 3
    for col_idx, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=_pretty_label(column_name))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_row = header_row + 1
    for index, row in enumerate(rows):
        mapping = _coerce_row_mapping(row)
        for col_idx, column_name in enumerate(columns, start=1):
            cell = sheet.cell(row=data_row, column=col_idx, value=_normalize_excel_value(mapping.get(column_name)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if index % 2 == 1:
                cell.fill = _ALT_FILL
        data_row += 1

    if not rows:
        empty_cell = sheet.cell(row=data_row, column=1, value="Sin datos")
        empty_cell.font = _EMPTY_FONT

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(end_col)}{max(header_row, data_row - 1)}"
    _autofit_columns(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _estimate_pdf_widths(columns: list[str], rows: list[dict[str, Any]], available_width: float) -> list[float]:
    sample_rows = rows[:75]
    widths: list[float] = []
    for column_name in columns:
        label = _pretty_label(column_name)
        max_len = len(label)
        for row in sample_rows:
            value = _normalize_pdf_value(_coerce_row_mapping(row).get(column_name))
            max_len = max(max_len, len(value))
        width = max(32.0, min(120.0, max_len * 4.3))
        widths.append(width)

    total = sum(widths)
    if not widths:
        return widths
    if total > available_width:
        scale = available_width / total
        widths = [max(28.0, width * scale) for width in widths]
    return widths


def _build_pdf(title: str, rows: list[dict[str, Any]]) -> bytes:
    columns = _collect_columns(rows)
    if not columns:
        columns = ["valor"]

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ExportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=18,
        textColor=colors.HexColor("#1F2937"),
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        "ExportMeta",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=4,
    )
    cell_style = ParagraphStyle(
        "ExportCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7,
        leading=8,
        textColor=colors.HexColor("#111827"),
    )
    header_style = ParagraphStyle(
        "ExportHeader",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=8,
        textColor=colors.white,
        alignment=1,
    )

    elements = [
        Paragraph(title or "Exportacion", title_style),
        Paragraph(f"Total de registros: {len(rows)}", meta_style),
        Spacer(1, 3 * mm),
    ]

    if not rows:
        elements.append(Paragraph("Sin datos para exportar.", meta_style))
        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    available_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
    col_widths = _estimate_pdf_widths(columns, rows, available_width)
    table_rows: list[list[Paragraph]] = [
        [Paragraph(_pretty_label(column_name), header_style) for column_name in columns]
    ]
    for row in rows:
        mapping = _coerce_row_mapping(row)
        table_rows.append(
            [Paragraph(_normalize_pdf_value(mapping.get(column_name)) or "&nbsp;", cell_style) for column_name in columns]
        )

    table = Table(table_rows, colWidths=col_widths, repeatRows=1)
    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("LEADING", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
    )
    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def build_table_export(title: str, rows: Iterable[Any], export_format: str) -> ExportFile:
    fmt = normalize_export_format(export_format)
    normalized_rows = [_coerce_row_mapping(row) for row in rows]

    if fmt == "xlsx":
        data = _build_xlsx(title, normalized_rows)
        return ExportFile(data=data, mimetype=_XLSX_MIME, filename="xlsx")
    if fmt == "pdf":
        data = _build_pdf(title, normalized_rows)
        return ExportFile(data=data, mimetype=_PDF_MIME, filename="pdf")

    raise ValueError(f"Formato de exportacion no soportado: {export_format}")


def build_attachment_response(data: bytes, *, filename: str, mimetype: str):
    return send_file(
        BytesIO(data),
        as_attachment=True,
        download_name=filename,
        mimetype=mimetype,
    )
