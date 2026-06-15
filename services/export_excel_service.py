from __future__ import annotations

import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
_TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
_SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
_SECTION_FONT = Font(bold=True, color="1F1F1F")
_HEADER_FILL = PatternFill("solid", fgColor="2F5597")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
_META_LABEL_FILL = PatternFill("solid", fgColor="E2E8F0")
_META_LABEL_FONT = Font(bold=True, color="1F2937")
_META_VALUE_FILL = PatternFill("solid", fgColor="FFFFFF")
_ERROR_FILL = PatternFill("solid", fgColor="FDE9E7")
_ERROR_FONT = Font(bold=True, color="9C0006")
_NOTE_FONT = Font(italic=True, color="6B7280")


def _normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "SI" if value else "NO"
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, datetime.timedelta):
        total_seconds = int(value.total_seconds())
        sign = "-" if total_seconds < 0 else ""
        total_seconds = abs(total_seconds)
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        if seconds:
            return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"
        return f"{sign}{hours:02d}:{minutes:02d}"
    return value


def _autofit(ws, *, min_width: int = 10, max_width: int = 45):
    for col_cells in ws.columns:
        values = [str(cell.value or "") for cell in col_cells]
        if not values:
            continue
        width = max(len(value) for value in values)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(width + 2, min_width),
            max_width,
        )


def _style_title_cell(cell):
    cell.fill = _TITLE_FILL
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_section_cell(cell):
    cell.fill = _SECTION_FILL
    cell.font = _SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_header_cell(cell):
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_meta_label(cell):
    cell.fill = _META_LABEL_FILL
    cell.font = _META_LABEL_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_meta_value(cell):
    cell.fill = _META_VALUE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_title(ws, title: str, end_col: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=1, value=title)
    _style_title_cell(cell)
    ws.row_dimensions[1].height = 22


def _write_section_title(ws, row: int, title: str, end_col: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1, value=title)
    _style_section_cell(cell)
    ws.row_dimensions[row].height = 18


def _write_header_row(ws, row: int, headers: list[str]):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        _style_header_cell(cell)


def _write_meta_pair(ws, row: int, label_col: int, value_col: int, label: str, value):
    label_cell = ws.cell(row=row, column=label_col, value=label)
    value_cell = ws.cell(row=row, column=value_col, value=_normalize_value(value))
    _style_meta_label(label_cell)
    _style_meta_value(value_cell)


def _write_table_block(ws, row: int, title: str, headers: list[str], rows: list[list]):
    _write_section_title(ws, row, title, max(1, len(headers)))
    _write_header_row(ws, row + 1, headers)
    current_row = row + 2
    if not rows:
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx, value="Sin datos" if col_idx == 1 else "")
            if col_idx == 1:
                cell.font = _NOTE_FONT
        return current_row + 2

    for index, values in enumerate(rows):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=_normalize_value(value))
            if index % 2 == 1:
                cell.fill = _ALT_FILL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        current_row += 1
    return current_row + 1


def _format_antiguedad(antiguedad: dict | None):
    if not antiguedad:
        return "-"
    return antiguedad.get("text") or "-"


def _format_name(empleado: dict | None):
    if not empleado:
        return "-"
    apellido = str(empleado.get("apellido") or "").strip()
    nombre = str(empleado.get("nombre") or "").strip()
    nombre_completo = f"{apellido} {nombre}".strip()
    return nombre_completo or "-"


def generar_planilla_fichadas_excel(
    *,
    planilla_rows: list[dict],
    empresa_sel: dict | None,
    sucursal_sel: dict | None,
    fecha: str,
    intervalo_minimo_fichadas: int,
    max_pares: int,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilla"

    total_cols = 2 + (max_pares * 2) + 1
    _write_title(ws, "Planilla diaria de fichadas", total_cols)

    _write_meta_pair(ws, 2, 1, 2, "Empresa", (empresa_sel or {}).get("razon_social") or "-")
    _write_meta_pair(ws, 2, 3, 4, "Sucursal", (sucursal_sel or {}).get("nombre") or "-")
    _write_meta_pair(ws, 3, 1, 2, "Fecha", fecha)
    _write_meta_pair(ws, 3, 3, 4, "Intervalo minimo", f"{intervalo_minimo_fichadas} min")
    _write_meta_pair(ws, 4, 1, 2, "Personas", len(planilla_rows))

    headers = ["Empleado", "DNI"]
    for idx in range(max_pares):
        headers.extend([f"Ingreso {idx + 1}", f"Egreso {idx + 1}"])
    headers.append("Errores")
    _write_header_row(ws, 5, headers)

    current_row = 6
    for row in planilla_rows:
        empleado = f"{row.get('apellido') or ''} {row.get('nombre') or ''}".strip()
        ws.cell(row=current_row, column=1, value=empleado or "-")
        ws.cell(row=current_row, column=2, value=row.get("dni") or "")

        pares = row.get("pares") or []
        for idx in range(max_pares):
            ingreso_col = 3 + (idx * 2)
            egreso_col = ingreso_col + 1
            par = pares[idx] if idx < len(pares) else None
            ws.cell(row=current_row, column=ingreso_col, value=(par or {}).get("ingreso") or "")
            ws.cell(row=current_row, column=egreso_col, value=(par or {}).get("egreso") or "")

        errores = " | ".join(row.get("errores") or [])
        error_cell = ws.cell(row=current_row, column=total_cols, value=errores)
        if errores:
            error_cell.fill = _ERROR_FILL
            error_cell.font = _ERROR_FONT

        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if current_row % 2 == 0 and col_idx != total_cols:
                cell.fill = _ALT_FILL

        current_row += 1

    ws.freeze_panes = "C6"
    ws.auto_filter.ref = f"A5:{get_column_letter(total_cols)}{max(5, current_row - 1)}"
    _autofit(ws, max_width=34)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_dashboard_empleado_excel(
    *,
    empleado: dict,
    desde: str,
    hasta: str,
    antiguedad: dict | None,
    stats: dict,
    legajo: dict,
    diaria_rows: list[dict],
    asistencia_rows: list[dict],
    eventos_periodo: list[dict],
) -> bytes:
    wb = Workbook()

    asistencia = stats.get("asistencia", {})

    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    _write_title(ws, "Dashboard por empleado", 5)

    row = 3
    row = _write_table_block(
        ws,
        row,
        "Datos del empleado",
        ["Campo", "Valor"],
        [
            ["Empleado", _format_name(empleado)],
            ["DNI", empleado.get("dni") or "-"],
            ["Legajo", empleado.get("legajo") or "-"],
            ["Empresa", empleado.get("empresa_nombre") or (f"Empresa #{empleado.get('empresa_id')}" if empleado.get("empresa_id") else "-")],
            ["Ingreso", empleado.get("fecha_ingreso") or "-"],
            ["Antiguedad", _format_antiguedad(antiguedad)],
            ["Periodo", f"{desde} a {hasta}"],
            ["Horario", stats.get("horario_desc") or "-"],
        ],
    )

    row = _write_table_block(
        ws,
        row,
        "KPIs de asistencia",
        ["Metrica", "Valor"],
        [
            ["Registros", asistencia.get("totales", {}).get("registros", 0)],
            ["OK / Puntual", asistencia.get("totales", {}).get("ok", 0)],
            ["Tardanzas", asistencia.get("totales", {}).get("tarde", 0)],
            ["Ausentes", asistencia.get("totales", {}).get("ausente", 0)],
            ["Salida anticipada", asistencia.get("totales", {}).get("salida_anticipada", 0)],
            ["Puntualidad %", asistencia.get("kpis", {}).get("puntualidad_pct", 0)],
            ["Ausentismo %", asistencia.get("kpis", {}).get("ausentismo_pct", 0)],
            ["No-show %", asistencia.get("kpis", {}).get("no_show_pct", 0)],
            ["Horas promedio", asistencia.get("kpis", {}).get("horas_promedio", 0)],
            ["Horas totales", asistencia.get("kpis", {}).get("horas_totales", 0)],
            ["GPS alertas", asistencia.get("kpis", {}).get("gps_incidencias", 0)],
            ["Racha OK", asistencia.get("racha_ok", 0)],
            ["Calidad fichada %", stats.get("calidad_fichada_pct", "")],
            ["Dias evaluados", stats.get("calidad_fichada_n", 0)],
            ["Justificaciones pendientes", asistencia.get("justificaciones", {}).get("pendientes", 0)],
            ["Vacaciones eventos", asistencia.get("vacaciones", {}).get("eventos", 0)],
            ["Vacaciones dias", asistencia.get("vacaciones", {}).get("dias", 0)],
        ],
    )

    row = _write_table_block(
        ws,
        row,
        "KPIs de legajo",
        ["Metrica", "Valor"],
        [
            ["Historico total", legajo.get("historico", {}).get("total", 0)],
            ["Historico vigentes", legajo.get("historico", {}).get("vigentes", 0)],
            ["Historico anulados", legajo.get("historico", {}).get("anulados", 0)],
            ["Periodo total", legajo.get("periodo", {}).get("total", 0)],
            ["Periodo vigentes", legajo.get("periodo", {}).get("vigentes", 0)],
            ["Periodo anulados", legajo.get("periodo", {}).get("anulados", 0)],
            ["Graves", legajo.get("periodo", {}).get("graves", 0)],
            ["Medios", legajo.get("periodo", {}).get("media", 0)],
            ["Leves", legajo.get("periodo", {}).get("leve", 0)],
            ["Tipos unicos", legajo.get("periodo", {}).get("tipos_unicos", 0)],
        ],
    )

    historial_rows = []
    for item in legajo.get("por_tipo_historico") or []:
        historial_rows.append(
            [
                item.get("label") or item.get("codigo") or "-",
                item.get("codigo") or "-",
                item.get("total") or 0,
                item.get("vigentes") or 0,
                item.get("ultima_fecha") or "",
            ]
        )
    row = _write_table_block(
        ws,
        row,
        "Eventos historicos por tipo",
        ["Tipo", "Codigo", "Total", "Vigentes", "Ultima fecha"],
        historial_rows,
    )

    ws.freeze_panes = "A4"
    _autofit(ws, max_width=34)

    # Fichadas
    ws = wb.create_sheet("Fichadas")
    _write_title(ws, "Fichadas del periodo", 9)
    _write_table_block(
        ws,
        3,
        "Detalle",
        [
            "Fecha",
            "Estado",
            "Hora entrada",
            "Hora salida",
            "Metodo entrada",
            "Metodo salida",
            "GPS entrada",
            "GPS salida",
            "Observaciones",
        ],
        [
            [
                row.get("fecha"),
                row.get("estado"),
                row.get("hora_entrada"),
                row.get("hora_salida"),
                row.get("metodo_entrada"),
                row.get("metodo_salida"),
                row.get("gps_ok_entrada"),
                row.get("gps_ok_salida"),
                row.get("observaciones"),
            ]
            for row in asistencia_rows
        ],
    )
    ws.freeze_panes = "A4"
    _autofit(ws, max_width=30)

    # Serie diaria
    ws = wb.create_sheet("Serie diaria")
    _write_title(ws, "Serie diaria", 9)
    _write_table_block(
        ws,
        3,
        "Detalle",
        ["Fecha", "Registros", "Horas", "Tramos", "Tardes", "Ausentes", "Salidas", "OK", "Estado"],
        [
            [
                row.get("fecha"),
                row.get("registros"),
                row.get("horas") or 0,
                row.get("tramos") or 0,
                row.get("tardes") or 0,
                row.get("ausentes") or 0,
                row.get("salidas") or 0,
                row.get("ok") or 0,
                row.get("estado") or "none",
            ]
            for row in diaria_rows
        ],
    )
    ws.freeze_panes = "A4"
    _autofit(ws, max_width=24)

    # Legajo
    ws = wb.create_sheet("Legajo")
    _write_title(ws, "Eventos de legajo", 10)
    _write_table_block(
        ws,
        3,
        "Periodo",
        [
            "Fecha evento",
            "Tipo",
            "Codigo",
            "Titulo",
            "Estado",
            "Severidad",
            "Fecha desde",
            "Fecha hasta",
            "Justificacion ID",
            "Descripcion",
        ],
        [
            [
                row.get("fecha_evento"),
                row.get("tipo_nombre") or row.get("tipo_codigo"),
                row.get("tipo_codigo"),
                row.get("titulo"),
                row.get("estado"),
                row.get("severidad"),
                row.get("fecha_desde"),
                row.get("fecha_hasta"),
                row.get("justificacion_id"),
                row.get("descripcion"),
            ]
            for row in eventos_periodo
        ],
    )
    ws.freeze_panes = "A4"
    _autofit(ws, max_width=32)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_historial_marcas_excel(*, rows: list[dict], filtros: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    _write_title(ws, "Historial de marcas", 8)

    filtros_rows = [
        ["Empresa", filtros.get("empresa_label") or (filtros.get("empresa_id") or "Todas")],
        ["Empleado", filtros.get("empleado_label") or (filtros.get("empleado_id") or "Todos")],
        ["Fecha desde", filtros.get("fecha_desde") or "-"],
        ["Fecha hasta", filtros.get("fecha_hasta") or "-"],
        ["Tipo de marca", filtros.get("tipo_marca") or "Todos"],
        ["Accion", filtros.get("accion") or "Todas"],
        ["Metodo", filtros.get("metodo") or "Todos"],
        ["GPS", "ok" if filtros.get("gps_ok") == 1 else "fuera" if filtros.get("gps_ok") == 0 else "Todos"],
        ["Busqueda", filtros.get("q") or "-"],
    ]
    row = _write_table_block(ws, 3, "Filtros", ["Campo", "Valor"], filtros_rows)

    detalle_rows = [
        [
            r.get("id"),
            r.get("empresa_nombre"),
            f"{r.get('apellido') or ''} {r.get('nombre') or ''}".strip(),
            r.get("dni"),
            r.get("fecha"),
            r.get("hora"),
            r.get("accion"),
            r.get("tipo_marca"),
            r.get("metodo"),
            r.get("gps_ok"),
            r.get("gps_distancia_m"),
            r.get("gps_tolerancia_m"),
            r.get("lat"),
            r.get("lon"),
            r.get("estado"),
            r.get("observaciones"),
            r.get("fecha_creacion"),
        ]
        for r in rows
    ]
    _write_table_block(
        ws,
        row,
        "Detalle",
        [
            "ID",
            "Empresa",
            "Empleado",
            "DNI",
            "Fecha",
            "Hora",
            "Accion",
            "Tipo",
            "Metodo",
            "GPS",
            "Dist (m)",
            "Tol (m)",
            "Lat",
            "Lon",
            "Estado",
            "Observaciones",
            "Fecha creacion",
        ],
        detalle_rows,
    )

    ws.freeze_panes = "A4"
    _autofit(ws, max_width=34)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_vacaciones_reporte_excel(*, rows: list[dict], totals: dict, filters: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    _write_title(ws, "Reporte de vacaciones", 10)

    filtros_rows = [
        ["Anio", filters.get("anio")],
        ["Sector", filters.get("sector_label") or (filters.get("sector_id") or "Todas")],
        ["Estado empleado", filters.get("activo_label") or filters.get("activo_raw") or "Todos"],
        ["Busqueda", filters.get("search") or "-"],
    ]
    row = _write_table_block(ws, 3, "Filtros", ["Campo", "Valor"], filtros_rows)

    resumen_rows = [
        ["Empleados", totals.get("empleados", 0)],
        ["Con error", totals.get("con_error", 0)],
        ["Dias corresponden", totals.get("dias_corresponden", 0)],
        ["Dias compensatorios", totals.get("dias_compensatorios", 0)],
        ["Dias ajustes", totals.get("dias_ajustes", 0)],
        ["Dias tomados", totals.get("dias_tomados", 0)],
        ["Dias pendientes", totals.get("dias_pendientes", 0)],
        ["Dias disponibles", totals.get("dias_disponibles", 0)],
        ["Dias disponibles con pendientes", totals.get("dias_disponibles_con_pendientes", 0)],
    ]
    row = _write_table_block(ws, row, "Totales", ["Metrica", "Valor"], resumen_rows)

    detalle_rows = [
        [
            r.get("empresa_nombre"),
            r.get("sector_nombre"),
            r.get("puesto_nombre"),
            f"{r.get('apellido') or ''} {r.get('nombre') or ''}".strip(),
            r.get("dni"),
            r.get("legajo"),
            r.get("estado"),
            r.get("fecha_ingreso"),
            r.get("antiguedad_al_31_12"),
            r.get("dias_base"),
            r.get("dias_ajustes"),
            r.get("dias_corresponden"),
            r.get("dias_compensatorios"),
            r.get("dias_tomados"),
            r.get("dias_pendientes_por_tomar"),
            r.get("dias_pendientes"),
            r.get("dias_disponibles_con_pendientes"),
            r.get("dias_trabajados_anio"),
            r.get("dias_habiles_anio"),
            "si" if r.get("calculo_proporcional") else "no",
            r.get("error"),
        ]
        for r in rows
    ]
    _write_table_block(
        ws,
        row,
        "Detalle",
        [
            "Empresa",
            "Area",
            "Puesto",
            "Empleado",
            "DNI",
            "Legajo",
            "Estado empleado",
            "Fecha ingreso",
            "Antiguedad al 31/12",
            "Dias base",
            "Dias ajustes",
            "Dias corresponden",
            "Dias compensatorios",
            "Dias tomados",
            "Pendientes por tomar",
            "Solicitudes pendientes",
            "Disponibles con pendientes",
            "Dias trabajados",
            "Dias habiles",
            "Proporcional",
            "Error",
        ],
        detalle_rows,
    )

    ws.freeze_panes = "A4"
    _autofit(ws, max_width=34)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_kpis_resultados_excel(
    *,
    empresa_label: str | None,
    sector_label: str | None,
    empleado_label: str | None,
    kpi_label: str | None,
    anio: int,
    mes: int,
    vista: dict,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    _write_title(ws, "Resultados KPI por empleado", 10)

    filtros_rows = [
        ["Empresa", empresa_label or "-"],
        ["Sector", sector_label or "-"],
        ["Empleado", empleado_label or "-"],
        ["KPI", kpi_label or "Todos"],
        ["Anio", anio],
        ["Mes", mes],
        ["Periodo", vista.get("month_label") or "-"],
    ]
    row = _write_table_block(ws, 3, "Filtros", ["Campo", "Valor"], filtros_rows)

    totales = vista.get("totales", {})
    resumen_totales_rows = [
        ["KPIs evaluados", totales.get("kpis", 0)],
        ["Dias con resultado", totales.get("dias_con_resultado", 0)],
        ["En objetivo", totales.get("en_objetivo", 0)],
        ["Alertas", totales.get("alerta", 0)],
        ["Sin datos", totales.get("sin_datos", 0)],
    ]
    row = _write_table_block(ws, row, "Totales", ["Metrica", "Valor"], resumen_totales_rows)

    resumen_rows = []
    for item in vista.get("resumen") or []:
        resumen_rows.append(
            [
                item.get("codigo"),
                item.get("nombre"),
                item.get("tipo_acumulacion"),
                item.get("objetivo_anual_label"),
                item.get("objetivo_mes_label"),
                item.get("resultado_mes"),
                item.get("resultado_anio"),
                item.get("semaforo_mes"),
                item.get("semaforo_anio"),
                item.get("mensaje_mes"),
                item.get("mensaje_anio"),
            ]
        )
    row = _write_table_block(
        ws,
        row,
        "Resumen mensual",
        [
            "Codigo",
            "Indicador",
            "Acumulacion",
            "Objetivo anual",
            "Objetivo mes",
            "Resultado mes",
            "Resultado anual",
            "Semaforo mes",
            "Semaforo anual",
            "Mensaje mes",
            "Mensaje anual",
        ],
        resumen_rows,
    )

    daily_rows = []
    for day in vista.get("daily_rows") or []:
        fecha = day.get("fecha")
        for entry in day.get("entries") or []:
            daily_rows.append(
                [
                    fecha,
                    entry.get("codigo"),
                    entry.get("nombre"),
                    entry.get("valor"),
                    entry.get("unidad"),
                    entry.get("objetivo_label"),
                    entry.get("semaforo"),
                    entry.get("mensaje"),
                ]
            )

    _write_table_block(
        ws,
        row,
        "Resultados diarios",
        ["Fecha", "Codigo", "Indicador", "Valor", "Unidad", "Objetivo", "Semaforo", "Mensaje"],
        daily_rows,
    )

    ws.freeze_panes = "A4"
    _autofit(ws, max_width=34)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_premios_concursos_resultados_excel(
    *,
    rows: list[dict],
    summary: dict,
    filters: dict,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    _write_title(ws, "Resultados de premios", 12)

    filtros_rows = [
        ["Empresa", filters.get("empresa_label") or "Todas"],
        ["Sector", filters.get("sector_label") or "Todos"],
        ["Empleado", filters.get("empleado_label") or "Todos"],
        ["Concurso", filters.get("concurso_label") or "Todos"],
        ["Anio", filters.get("periodo_year") or "-"],
        ["Mes", filters.get("periodo_month") or "-"],
        ["Busqueda", filters.get("search") or "-"],
    ]
    row = _write_table_block(ws, 3, "Filtros", ["Campo", "Valor"], filtros_rows)

    resumen_rows = [
        ["Resultados", summary.get("total", 0)],
        ["Habilitados", summary.get("habilitados", 0)],
        ["Completados", summary.get("completados", 0)],
        ["En progreso", summary.get("en_progreso", 0)],
        ["Pendientes", summary.get("pendientes", 0)],
        ["Excluidos", summary.get("excluidos", 0)],
        ["Empleados", summary.get("empleados", 0)],
        ["Concursos", summary.get("concursos", 0)],
        ["Primeros puestos", summary.get("primeros", 0)],
        ["Podios", summary.get("podios", 0)],
    ]
    row = _write_table_block(ws, row, "Resumen", ["Metrica", "Valor"], resumen_rows)

    detalle_rows = []
    for r in rows:
        detalle_rows.append(
            [
                r.get("id"),
                r.get("empresa_nombre"),
                r.get("periodo_label"),
                r.get("legajo") or r.get("legajo_snapshot") or "",
                r.get("dni"),
                r.get("apellido"),
                r.get("nombre"),
                r.get("concurso_codigo") or r.get("concurso_codigo_snapshot") or "",
                r.get("concurso_nombre") or r.get("concurso_nombre_snapshot") or "",
                r.get("concurso_alcance"),
                r.get("sector_nombre") or "",
                r.get("ranking"),
                r.get("observaciones"),
            ]
        )

    _write_table_block(
        ws,
        row,
        "Detalle",
        [
            "ID",
            "Empresa",
            "Periodo",
            "Legajo",
            "DNI",
            "Apellido",
            "Nombre",
            "Codigo concurso",
            "Concurso",
            "Alcance",
            "Sector",
            "Ranking",
            "Observaciones",
        ],
        detalle_rows,
    )

    ws.freeze_panes = "A4"
    _autofit(ws, max_width=34)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_trivia_resultados_excel(
    *,
    trivia: dict,
    rows: list[dict],
    summary: dict,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    _write_title(ws, "Resultados de trivia", 12)

    filtros_rows = [
        ["Trivia", trivia.get("titulo") or "-"],
        ["Estado", trivia.get("estado") or "-"],
        ["ID", trivia.get("id") or "-"],
    ]
    row = _write_table_block(ws, 3, "Trivia", ["Campo", "Valor"], filtros_rows)

    resumen_rows = [
        ["Habilitados", summary.get("habilitados", 0)],
        ["Resultados", summary.get("resultados", 0)],
        ["Completados", summary.get("completados", 0)],
        ["En progreso", summary.get("en_progreso", 0)],
        ["Pendientes", summary.get("pendientes", 0)],
        ["Excluidos", summary.get("excluidos", 0)],
    ]
    row = _write_table_block(ws, row, "Resumen", ["Metrica", "Valor"], resumen_rows)

    detalle_rows = []
    for r in rows:
        detalle_rows.append(
            [
                trivia.get("id"),
                trivia.get("titulo"),
                trivia.get("estado"),
                r.get("empleado_id"),
                r.get("empleado_legajo") or "",
                r.get("empleado_dni") or "",
                r.get("empleado_apellido") or "",
                r.get("empleado_nombre") or "",
                r.get("sector_nombre") or "",
                r.get("estado_admin") or "",
                r.get("posicion_calculada") or "",
                r.get("puntos_total") if r.get("resultado_id") else "",
                r.get("correctas") if r.get("resultado_id") else "",
                r.get("incorrectas") if r.get("resultado_id") else "",
                r.get("tiempo_total_segundos") if r.get("resultado_id") else "",
                _normalize_value(r.get("fecha_inicio_participacion")),
                _normalize_value(r.get("fecha_finalizacion")),
                r.get("exclusion_motivo") or "",
            ]
        )

    _write_table_block(
        ws,
        row,
        "Detalle",
        [
            "Trivia ID",
            "Trivia",
            "Estado trivia",
            "Empleado ID",
            "Legajo",
            "DNI",
            "Apellido",
            "Nombre",
            "Sector",
            "Estado",
            "Posicion",
            "Puntos",
            "Correctas",
            "Incorrectas",
            "Tiempo segundos",
            "Inicio participacion",
            "Fin participacion",
            "Motivo exclusion",
        ],
        detalle_rows,
    )

    ws.freeze_panes = "A4"
    _autofit(ws, max_width=34)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
