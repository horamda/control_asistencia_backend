"""
Genera el template Excel para importación masiva de empleados.

Hoja 1 – Plantilla: encabezados + fila de ejemplo (esta hoja se exporta como CSV).
Hoja 2 – Empresas: id y razón social disponibles.
Hoja 3 – Sucursales: nombre exacto necesario en el CSV, agrupado por empresa.
Hoja 4 – Sectores: nombre exacto necesario en el CSV, agrupado por empresa.
Hoja 5 – Puestos: nombre exacto necesario en el CSV, agrupado por empresa.
Hoja 6 – Localidades: código postal y localidad disponibles.
Hoja 7 – Valores válidos: enumerados aceptados por cada campo.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from repositories.empresa_repository import get_all as get_empresas
from repositories.sucursal_repository import get_all as get_sucursales
from repositories.sector_repository import get_all as get_sectores
from repositories.puesto_repository import get_all as get_puestos
from repositories.localidad_repository import get_all as get_localidades


# ── Estilos ──────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_REF_FILL = PatternFill("solid", fgColor="2E75B6")
_REF_FONT = Font(bold=True, color="FFFFFF")
_REQUIRED_FILL = PatternFill("solid", fgColor="FFE699")
_NOTE_FILL = PatternFill("solid", fgColor="E2EFDA")
_NOTE_FONT = Font(italic=True, color="375623")


def _set_header_row(ws, cols: list[str], fill=None, font=None, row: int = 1):
    """Escribe una fila de encabezado con estilo."""
    fill = fill or _HEADER_FILL
    font = font or _HEADER_FONT
    for col_idx, text in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws, min_width: int = 10, max_width: int = 45):
    """Ajusta el ancho de columnas al contenido."""
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value or "")) for c in col_cells),
            default=min_width,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(length + 3, min_width), max_width
        )


# ── Hoja 1: Plantilla ────────────────────────────────────────────────────────

_COLUMNS = [
    ("legajo",           "Legajo",           True,  "001"),
    ("dni",              "DNI",              True,  "12345678"),
    ("cuil",             "CUIL",             False, "20-12345678-5"),
    ("apellido",         "Apellido",         True,  "Perez"),
    ("nombre",           "Nombre",           True,  "Juan"),
    ("sexo",             "Sexo",             False, "masculino"),
    ("fecha_nacimiento", "Fecha nacimiento", False, "1990-05-20"),
    ("email",            "Email",            False, "juan@empresa.com"),
    ("telefono",         "Telefono",         False, "1123456789"),
    ("direccion",        "Direccion",        False, "Av. Corrientes 1234"),
    ("codigo_postal",    "Codigo postal",    False, "C1043AAZ"),
    ("fecha_ingreso",    "Fecha ingreso",    False, "2024-01-01"),
    ("tipo_contrato",    "Tipo contrato",    False, "efectivo"),
    ("modalidad",        "Modalidad",        False, "presencial"),
    ("categoria",        "Categoria",        False, "Categoria A"),
    ("obra_social",      "Obra social",      False, "OSDE"),
    ("cod_chess_erp",    "Cod. Chess ERP",   False, "999"),
    ("banco",            "Banco",            False, "Banco Nacion"),
    ("cbu",              "CBU",              False, "0110000900000012345678"),
    ("numero_emergencia","Nro. emergencia",  False, "1158887766"),
    ("estado",           "Estado",           False, "activo"),
    ("sucursal_nombre",  "Sucursal nombre",  False, "Sucursal Centro"),
    ("sector_nombre",    "Sector nombre",    False, "Administracion"),
    ("puesto_nombre",    "Puesto nombre",    False, "Analista"),
    ("password",         "Password",         False, ""),
]


def _build_plantilla(wb: Workbook):
    ws = wb.active
    ws.title = "Plantilla"

    # Nota instructiva en fila 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))
    note_cell = ws.cell(row=1, column=1)
    note_cell.value = (
        "INSTRUCCIONES: Complete los datos en esta hoja (fila 4 en adelante). "
        "Columnas en amarillo son OBLIGATORIAS. "
        "Consulte las hojas de referencia para valores exactos de sucursal, sector, puesto, localidad y enumerados. "
        "Cuando termine, exporte ESTA HOJA como CSV (UTF-8) y carguela en el sistema."
    )
    note_cell.fill = _NOTE_FILL
    note_cell.font = _NOTE_FONT
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 40

    # Fila 2: nombre de campo (clave CSV)
    _set_header_row(ws, [c[0] for c in _COLUMNS], row=2)

    # Fila 3: etiqueta legible + marcador de obligatorio
    for col_idx, (key, label, required, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=f"{label}{'  *' if required else ''}")
        if required:
            cell.fill = _REQUIRED_FILL
            cell.font = Font(bold=True, color="7B3F00")
        else:
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.font = Font(color="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # Fila 4: fila de ejemplo
    for col_idx, (_, _, _, ejemplo) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=ejemplo)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.font = Font(italic=True, color="595959")

    # Freeze header
    ws.freeze_panes = "A5"
    _autofit(ws)


# ── Hojas de referencia ───────────────────────────────────────────────────────

def _build_empresas(wb: Workbook):
    ws = wb.create_sheet("Empresas")
    _set_header_row(ws, ["ID", "Razon social"], fill=_REF_FILL, font=_REF_FONT)
    for e in get_empresas(include_inactive=False):
        ws.append([e.get("id"), e.get("razon_social")])
    _autofit(ws)


def _build_sucursales(wb: Workbook):
    ws = wb.create_sheet("Sucursales")
    _set_header_row(ws, ["Empresa", "Nombre (usar en CSV)", "ID"], fill=_REF_FILL, font=_REF_FONT)
    for s in get_sucursales(include_inactive=False):
        ws.append([s.get("empresa_nombre"), s.get("nombre"), s.get("id")])
    _autofit(ws)


def _build_sectores(wb: Workbook):
    ws = wb.create_sheet("Sectores")
    _set_header_row(ws, ["Empresa", "Nombre (usar en CSV)", "ID"], fill=_REF_FILL, font=_REF_FONT)
    for s in get_sectores(include_inactive=False):
        ws.append([s.get("empresa_nombre"), s.get("nombre"), s.get("id")])
    _autofit(ws)


def _build_puestos(wb: Workbook):
    ws = wb.create_sheet("Puestos")
    _set_header_row(ws, ["Empresa", "Nombre (usar en CSV)", "ID"], fill=_REF_FILL, font=_REF_FONT)
    for p in get_puestos(include_inactive=False):
        ws.append([p.get("empresa_nombre"), p.get("nombre"), p.get("id")])
    _autofit(ws)


def _build_localidades(wb: Workbook):
    ws = wb.create_sheet("Localidades")
    _set_header_row(ws, ["Codigo postal (usar en CSV)", "Localidad", "Provincia"], fill=_REF_FILL, font=_REF_FONT)
    for loc in get_localidades():
        ws.append([loc.get("codigo_postal"), loc.get("localidad"), loc.get("provincia")])
    _autofit(ws)


def _build_valores_validos(wb: Workbook):
    ws = wb.create_sheet("Valores validos")
    _set_header_row(ws, ["Campo CSV", "Valores aceptados", "Notas"], fill=_REF_FILL, font=_REF_FONT)
    rows = [
        ("sexo",           "masculino | femenino | no_binario | no_informa",       "Default: no_informa"),
        ("estado",         "activo | inactivo | suspendido",                        "Default: activo"),
        ("tipo_contrato",  "efectivo | temporal | pasantia | otro",                 "Opcional"),
        ("modalidad",      "presencial | remoto | hibrido",                         "Default: presencial"),
        ("password",       "(cualquier texto)",                                      "Si se omite, se usa el DNI como contrasena inicial"),
        ("fecha_*",        "YYYY-MM-DD",                                            "Ejemplo: 2024-01-15"),
        ("cbu",            "22 digitos numericos",                                  ""),
        ("cod_chess_erp",  "Numero entero",                                         ""),
    ]
    for r in rows:
        ws.append(list(r))
    _autofit(ws)


# ── Función pública ───────────────────────────────────────────────────────────

def generar_template_excel() -> bytes:
    """
    Genera el workbook Excel y devuelve los bytes listos para servir.
    """
    wb = Workbook()

    _build_plantilla(wb)
    _build_empresas(wb)
    _build_sucursales(wb)
    _build_sectores(wb)
    _build_puestos(wb)
    _build_localidades(wb)
    _build_valores_validos(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
