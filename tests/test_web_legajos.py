import datetime
import io

from openpyxl import load_workbook

import app as app_module
import web.auth.decorators as auth_decorators
import web.legajos.legajos_routes as legajos_routes
import web.legajos.legajo_tipos_evento_routes as legajo_tipos_evento_routes


def _build_client(monkeypatch):
    monkeypatch.setattr(app_module, "init_db", lambda: None)
    app = app_module.create_app()
    app.config["TESTING"] = True
    app.config["WTF_CSRF_ENABLED"] = False
    return app.test_client()


def _login_session(client):
    with client.session_transaction() as sess:
        sess["user_id"] = 11


def test_legajos_listado_eventos(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    with client.session_transaction() as sess:
        sess["user_role"] = "admin"
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        legajos_routes,
        "get_eventos_page",
        lambda **kwargs: (
            [
                {
                    "id": 44,
                    "empresa_id": 3,
                    "empresa_nombre": "Empresa A",
                    "empleado_id": 7,
                    "empleado_apellido": "Perez",
                    "empleado_nombre": "Ana",
                    "empleado_legajo": "L-99",
                    "empleado_dni": "30123456",
                    "empleado_foto": None,
                    "tipo_nombre": "Certificado medico",
                    "fecha_evento": "2026-03-05",
                    "titulo": "Certificado",
                    "estado": "vigente",
                }
            ],
            1,
        ),
    )
    monkeypatch.setattr(legajos_routes, "get_empresas", lambda include_inactive=True: [{"id": 3, "razon_social": "Empresa A"}])
    monkeypatch.setattr(
        legajos_routes,
        "get_empleados",
        lambda include_inactive=True: [{"id": 7, "apellido": "Perez", "nombre": "Ana", "dni": "30123456"}],
    )
    monkeypatch.setattr(legajos_routes, "get_tipos_evento", lambda include_inactive=True: [{"id": 1, "nombre": "Certificado medico"}])
    monkeypatch.setattr(legajos_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(legajos_routes, "get_sectores", lambda include_inactive=True: [])

    resp = client.get("/legajos/eventos/?empresa_id=3")
    assert resp.status_code == 200
    assert b"Certificado medico" in resp.data
    assert b"/legajos/empleado/7" in resp.data


def test_legajos_listado_eventos_aplica_fechas_y_severidad(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    with client.session_transaction() as sess:
        sess["user_role"] = "admin"
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    captured = {}
    monkeypatch.setattr(
        legajos_routes,
        "get_eventos_page",
        lambda **kwargs: captured.update(kwargs) or ([], 0),
    )
    monkeypatch.setattr(legajos_routes, "get_empresas", lambda include_inactive=True: [])
    monkeypatch.setattr(legajos_routes, "get_empleados", lambda include_inactive=True: [])
    monkeypatch.setattr(legajos_routes, "get_tipos_evento", lambda include_inactive=True: [])
    monkeypatch.setattr(legajos_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(legajos_routes, "get_sectores", lambda include_inactive=True: [])

    resp = client.get("/legajos/eventos/?severidad=grave&fecha_desde=2026-07-01&fecha_hasta=2026-07-31&q=L-99")

    assert resp.status_code == 200
    assert captured["severidad"] == "grave"
    assert captured["fecha_desde"] == "2026-07-01"
    assert captured["fecha_hasta"] == "2026-07-31"
    assert captured["search"] == "L-99"


def test_legajos_listado_muestra_foto_y_fallback(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        legajos_routes,
        "_get_empleados_page",
        lambda *args, **kwargs: ([
            {
                "id": 7,
                "empresa_id": 3,
                "empresa_nombre": "Empresa A",
                "legajo": "L-99",
                "dni": "30123456",
                "nombre": "Ana",
                "apellido": "Perez",
                "activo": 1,
                "foto": "https://cdn.example.com/fotos/30123456.jpg",
                "legajo_eventos_total": 2,
                "legajo_eventos_vigentes": 1,
            },
            {
                "id": 8,
                "empresa_id": 3,
                "empresa_nombre": "Empresa A",
                "legajo": "L-100",
                "dni": "30123457",
                "nombre": "Juan",
                "apellido": "Lopez",
                "activo": 1,
                "foto": None,
                "legajo_eventos_total": 0,
                "legajo_eventos_vigentes": 0,
            },
        ], 2),
    )
    monkeypatch.setattr(legajos_routes, "get_empresas", lambda include_inactive=True: [{"id": 3, "razon_social": "Empresa A"}])
    monkeypatch.setattr(legajos_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(legajos_routes, "get_sectores", lambda include_inactive=True: [])

    resp = client.get("/legajos/")
    assert resp.status_code == 200
    assert b"https://cdn.example.com/fotos/30123456.jpg" in resp.data
    assert b"img/empleado-default.svg" in resp.data


def test_legajos_listado_empleados_aplica_filtros(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    captured = {}

    def fake_page(*args, **kwargs):
        captured.update(kwargs)
        return ([], 0)

    monkeypatch.setattr(legajos_routes, "_get_empleados_page", fake_page)
    monkeypatch.setattr(legajos_routes, "get_empresas", lambda include_inactive=True: [])
    monkeypatch.setattr(legajos_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(legajos_routes, "get_sectores", lambda include_inactive=True: [])

    resp = client.get(
        "/legajos/?q=L-99&empresa_id=3&sucursal_id=4&sector_id=5&activo=all"
        "&requiere_control_asistencia=0&legajo_eventos=vigentes&per=50"
    )

    assert resp.status_code == 200
    assert captured["search"] == "L-99"
    assert captured["empresa_id"] == 3
    assert captured["sucursal_id"] == 4
    assert captured["sector_id"] == 5
    assert captured["activo"] is None
    assert captured["requiere_control_asistencia"] == 0
    assert captured["legajo_eventos"] == "vigentes"


def test_legajo_empleado_muestra_fallback_si_no_tiene_foto(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        legajos_routes,
        "get_empleado_by_id",
        lambda emp_id: {
            "id": emp_id,
            "empresa_id": 3,
            "legajo": "L-99",
            "dni": "30123456",
            "nombre": "Ana",
            "apellido": "Perez",
            "foto": None,
        },
    )
    monkeypatch.setattr(legajos_routes, "get_eventos_by_empleado", lambda emp_id, include_anulados=True: [])
    monkeypatch.setattr(legajos_routes, "get_tipos_evento", lambda include_inactive=False: [])

    resp = client.get("/legajos/empleado/7")
    assert resp.status_code == 200
    assert b"img/empleado-default.svg" in resp.data


def test_dashboard_empleado_suma_tramos_y_muestra_antiguedad(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)

    empleado = {
        "id": 7,
        "empresa_id": 3,
        "empresa_nombre": "Empresa A",
        "legajo": "L-99",
        "dni": "30123456",
        "nombre": "Ana",
        "apellido": "Perez",
        "fecha_ingreso": "2020-01-15",
        "foto": None,
    }
    filas = [
        {
            "fecha": "2026-06-01",
            "hora_entrada": datetime.timedelta(hours=7),
            "hora_salida": datetime.timedelta(hours=12),
            "estado": "ok",
            "gps_ok_entrada": 1,
            "gps_ok_salida": 1,
            "metodo_entrada": "manual",
            "metodo_salida": "manual",
        },
        {
            "fecha": "2026-06-01",
            "hora_entrada": datetime.timedelta(hours=14),
            "hora_salida": datetime.timedelta(hours=17),
            "estado": "ok",
            "gps_ok_entrada": 1,
            "gps_ok_salida": 1,
            "metodo_entrada": "manual",
            "metodo_salida": "manual",
        },
    ]

    monkeypatch.setattr(legajos_routes, "_get_empleados_page", lambda *args, **kwargs: ([empleado], 1))
    monkeypatch.setattr(legajos_routes, "get_empleado_by_id", lambda emp_id: dict(empleado))
    monkeypatch.setattr(legajos_routes, "_get_asistencias_page", lambda *args, **kwargs: (list(filas), len(filas)))
    monkeypatch.setattr(legajos_routes, "_get_justificaciones_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(legajos_routes, "_get_vacaciones_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(
        legajos_routes,
        "calcular_resumen_vacaciones",
        lambda empleado_id, anio: {
            "anio": anio,
            "vacaciones": {
                "dias_base": 14,
                "dias_corresponden": 14,
                "dias_tomados": 4,
                "dias_disponibles": 10,
                "dias_disponibles_con_pendientes": 8,
                "dias_pendientes": 2,
                "dias_compensatorios": 0,
                "dias_ajustes": 0,
                "aplica_control_proporcional": False,
                "calculo_proporcional": False,
                "antiguedad_al_31_12": 6,
                "desglose_corresponde": [{"concepto": "Base LCT", "dias": 14}],
                "fecha_evaluacion_trabajo": "2026-12-31",
            },
        },
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_eventos_by_empleado",
        lambda emp_id, include_anulados=True: [
            {
                "id": 1,
                "estado": "vigente",
                "fecha_evento": "2026-05-10",
                "tipo_id": 10,
                "tipo_nombre": "Vacaciones",
                "tipo_codigo": "VAC",
                "severidad": "leve",
            }
        ],
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_conteo_por_tipo_for_empleado",
        lambda emp_id: [
            {"tipo_id": 10, "codigo": "VAC", "nombre": "Vacaciones", "total": 1, "vigentes": 1, "ultima_fecha": datetime.date(2026, 5, 10)},
            {"tipo_id": 11, "codigo": "AMO", "nombre": "Amonestacion", "total": 2, "vigentes": 1, "ultima_fecha": datetime.date(2026, 5, 18)},
        ],
    )

    resp = client.get("/legajos/dashboard-empleado?empleado_id=7&desde=2026-05-01&hasta=2026-06-30&periodo=custom")
    assert resp.status_code == 200
    assert b"6 anios, 5 meses, 15 dias" in resp.data
    assert b"8.0h" in resp.data
    assert b"2 tramos" in resp.data
    assert b"Vacaciones 2026" in resp.data
    assert b"4 tomados / 10 disponibles" in resp.data
    assert b"Vacaciones anuales" in resp.data
    assert b"Eventos de legajo por tipo (historico)" in resp.data
    assert b"Vacaciones" in resp.data
    assert b"Amonestacion" in resp.data


def test_dashboard_empleado_export_xlsx_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)

    empleado = {
        "id": 7,
        "empresa_id": 3,
        "empresa_nombre": "Empresa A",
        "legajo": "L-99",
        "dni": "30123456",
        "nombre": "Ana",
        "apellido": "Perez",
        "fecha_ingreso": "2020-01-15",
        "foto": None,
    }
    filas = [
        {
            "fecha": "2026-06-01",
            "hora_entrada": datetime.timedelta(hours=7),
            "hora_salida": datetime.timedelta(hours=12),
            "estado": "ok",
            "gps_ok_entrada": 1,
            "gps_ok_salida": 1,
            "metodo_entrada": "manual",
            "metodo_salida": "manual",
            "observaciones": "Sin novedad",
        },
        {
            "fecha": "2026-06-01",
            "hora_entrada": datetime.timedelta(hours=14),
            "hora_salida": datetime.timedelta(hours=17),
            "estado": "ok",
            "gps_ok_entrada": 1,
            "gps_ok_salida": 1,
            "metodo_entrada": "manual",
            "metodo_salida": "manual",
            "observaciones": "Sin novedad",
        },
    ]

    monkeypatch.setattr(legajos_routes, "_get_empleados_page", lambda *args, **kwargs: ([empleado], 1))
    monkeypatch.setattr(legajos_routes, "get_empleado_by_id", lambda emp_id: dict(empleado))
    monkeypatch.setattr(legajos_routes, "_get_asistencias_page", lambda *args, **kwargs: (list(filas), len(filas)))
    monkeypatch.setattr(legajos_routes, "_get_justificaciones_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(legajos_routes, "_get_vacaciones_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(
        legajos_routes,
        "calcular_resumen_vacaciones",
        lambda empleado_id, anio: {
            "anio": anio,
            "vacaciones": {
                "dias_base": 14,
                "dias_corresponden": 14,
                "dias_tomados": 4,
                "dias_disponibles": 10,
                "dias_disponibles_con_pendientes": 8,
                "dias_pendientes": 2,
                "dias_compensatorios": 0,
                "dias_ajustes": 0,
                "aplica_control_proporcional": False,
                "calculo_proporcional": False,
                "antiguedad_al_31_12": 6,
                "desglose_corresponde": [{"concepto": "Base LCT", "dias": 14}],
                "fecha_evaluacion_trabajo": "2026-12-31",
            },
        },
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_eventos_by_empleado",
        lambda emp_id, include_anulados=True: [
            {
                "id": 1,
                "estado": "vigente",
                "fecha_evento": "2026-05-10",
                "tipo_id": 10,
                "tipo_nombre": "Vacaciones",
                "tipo_codigo": "VAC",
                "severidad": "leve",
                "fecha_desde": "2026-05-09",
                "fecha_hasta": "2026-05-11",
                "titulo": "Vacaciones de invierno",
                "descripcion": "Descanso programado",
            }
        ],
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_conteo_por_tipo_for_empleado",
        lambda emp_id: [
            {"tipo_id": 10, "codigo": "VAC", "nombre": "Vacaciones", "total": 1, "vigentes": 1, "ultima_fecha": datetime.date(2026, 5, 10)},
            {"tipo_id": 11, "codigo": "AMO", "nombre": "Amonestacion", "total": 2, "vigentes": 1, "ultima_fecha": datetime.date(2026, 5, 18)},
        ],
    )

    resp = client.get("/legajos/dashboard-empleado/export.xls?empleado_id=7&desde=2026-05-01&hasta=2026-06-01&periodo=custom")
    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers["Content-Type"]
    assert "resumen_Perez_Ana_2026-05-01_2026-06-01.xlsx" in resp.headers["Content-Disposition"]

    wb = load_workbook(io.BytesIO(resp.data), data_only=True)
    assert wb.sheetnames == ["Resumen", "Fichadas", "Serie diaria", "Legajo"]

    resumen = wb["Resumen"]
    assert resumen["A1"].value == "Resumen por empleado"
    assert resumen["A3"].value == "Datos del empleado"
    assert resumen["A5"].value == "Empleado"
    assert resumen["B5"].value == "Perez Ana"
    assert any(cell.value == "Vacaciones" for row in resumen.iter_rows() for cell in row)
    assert any(cell.value == "Corresponden" for row in resumen.iter_rows() for cell in row)
    assert any(cell.value == "Disponibles con pendientes" for row in resumen.iter_rows() for cell in row)
    assert any(cell.value == "Amonestacion" for row in resumen.iter_rows() for cell in row)

    fichadas = wb["Fichadas"]
    assert fichadas["A5"].value == "2026-06-01"
    assert fichadas["C5"].value == "07:00"
    assert fichadas["D5"].value == "12:00"

    serie = wb["Serie diaria"]
    assert serie["A5"].value == "2026-06-01"
    assert serie["C5"].value == 8.0
    assert serie["D5"].value == 2

    legajo = wb["Legajo"]
    assert legajo["A5"].value == "2026-05-10"
    assert legajo["B5"].value == "Vacaciones"
    assert legajo["D5"].value == "Vacaciones de invierno"


def test_dashboard_empleado_print_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)

    empleado = {
        "id": 7,
        "empresa_id": 3,
        "empresa_nombre": "Empresa A",
        "legajo": "L-99",
        "dni": "30123456",
        "nombre": "Ana",
        "apellido": "Perez",
        "fecha_ingreso": "2020-01-15",
        "foto": None,
    }
    filas = [
        {
            "fecha": "2026-06-01",
            "hora_entrada": datetime.timedelta(hours=7),
            "hora_salida": datetime.timedelta(hours=12),
            "estado": "ok",
            "gps_ok_entrada": 1,
            "gps_ok_salida": 1,
            "metodo_entrada": "manual",
            "metodo_salida": "manual",
        }
    ]

    monkeypatch.setattr(legajos_routes, "_get_empleados_page", lambda *args, **kwargs: ([empleado], 1))
    monkeypatch.setattr(legajos_routes, "get_empleado_by_id", lambda emp_id: dict(empleado))
    monkeypatch.setattr(legajos_routes, "_get_asistencias_page", lambda *args, **kwargs: (list(filas), len(filas)))
    monkeypatch.setattr(legajos_routes, "_get_justificaciones_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(legajos_routes, "_get_vacaciones_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(
        legajos_routes,
        "calcular_resumen_vacaciones",
        lambda empleado_id, anio: {
            "anio": anio,
            "vacaciones": {
                "dias_base": 14,
                "dias_corresponden": 14,
                "dias_tomados": 4,
                "dias_disponibles": 10,
                "dias_disponibles_con_pendientes": 8,
                "dias_pendientes": 2,
                "dias_compensatorios": 0,
                "dias_ajustes": 0,
                "aplica_control_proporcional": False,
                "calculo_proporcional": False,
                "antiguedad_al_31_12": 6,
                "desglose_corresponde": [{"concepto": "Base LCT", "dias": 14}],
                "fecha_evaluacion_trabajo": "2026-12-31",
            },
        },
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_eventos_by_empleado",
        lambda emp_id, include_anulados=True: [
            {
                "id": 1,
                "estado": "vigente",
                "fecha_evento": "2026-05-10",
                "tipo_id": 10,
                "tipo_nombre": "Vacaciones",
                "tipo_codigo": "VAC",
                "severidad": "leve",
            }
        ],
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_conteo_por_tipo_for_empleado",
        lambda emp_id: [
            {"tipo_id": 10, "codigo": "VAC", "nombre": "Vacaciones", "total": 1, "vigentes": 1, "ultima_fecha": datetime.date(2026, 5, 10)},
        ],
    )

    resp = client.get("/legajos/dashboard-empleado/print?empleado_id=7&desde=2026-05-01&hasta=2026-06-30&periodo=custom&auto_print=1")
    assert resp.status_code == 200
    assert b"Vacaciones anuales" in resp.data
    assert b"Tomados aprobados" in resp.data


def test_legajos_crear_evento_con_adjunto(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajos_routes, "log_audit", lambda *args, **kwargs: None)

    monkeypatch.setattr(
        legajos_routes,
        "get_empleado_by_id",
        lambda emp_id: {
            "id": emp_id,
            "empresa_id": 3,
            "legajo": "L-99",
            "dni": "30123456",
            "nombre": "Ana",
            "apellido": "Perez",
        },
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_tipo_evento_by_id",
        lambda tipo_id: {
            "id": tipo_id,
            "activo": 1,
            "requiere_rango_fechas": 0,
        },
    )

    captured = {}

    def _fake_create_evento(data):
        captured["evento_data"] = dict(data)
        return 77

    def _fake_save(file_storage, **kwargs):
        captured["saved_kwargs"] = dict(kwargs)
        captured["saved_filename"] = file_storage.filename
        return {
            "nombre_original": "certificado.pdf",
            "mime_type": "application/pdf",
            "extension": "pdf",
            "tamano_bytes": 1234,
            "sha256": "a" * 64,
            "storage_backend": "local",
            "storage_ruta": "uploads/legajos/empresa_3/empleado_7/evento_77/fake.pdf",
        }

    def _fake_create_adjunto(data):
        captured["adjunto_data"] = dict(data)
        return 88

    monkeypatch.setattr(legajos_routes, "create_evento", _fake_create_evento)
    monkeypatch.setattr(legajos_routes, "save_legajo_attachment_local", _fake_save)
    monkeypatch.setattr(legajos_routes, "create_adjunto", _fake_create_adjunto)

    resp = client.post(
        "/legajos/empleado/7/eventos",
        data={
            "tipo_id": "1",
            "fecha_evento": "2026-03-05",
            "descripcion": "Certificado por enfermedad",
            "adjuntos": (io.BytesIO(b"%PDF-1.4"), "certificado.pdf"),
        },
        content_type="multipart/form-data",
        follow_redirects=False,
    )

    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/legajos/empleado/7")
    assert captured["evento_data"]["empresa_id"] == 3
    assert captured["evento_data"]["empleado_id"] == 7
    assert captured["saved_filename"] == "certificado.pdf"
    assert captured["saved_kwargs"]["empresa_id"] == 3
    assert captured["saved_kwargs"]["empleado_id"] == 7
    assert captured["saved_kwargs"]["evento_id"] == 77
    assert captured["adjunto_data"]["evento_id"] == 77
    assert captured["adjunto_data"]["empresa_id"] == 3
    assert captured["adjunto_data"]["empleado_id"] == 7


def test_legajos_editar_evento(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajos_routes, "log_audit", lambda *args, **kwargs: None)
    monkeypatch.setattr(
        legajos_routes,
        "get_empleado_by_id",
        lambda emp_id: {
            "id": emp_id,
            "empresa_id": 3,
            "legajo": "L-99",
            "dni": "30123456",
            "nombre": "Ana",
            "apellido": "Perez",
        },
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_evento_by_id",
        lambda evento_id: {
            "id": evento_id,
            "empleado_id": 7,
            "tipo_id": 1,
            "fecha_evento": None,
            "fecha_desde": None,
            "fecha_hasta": None,
            "titulo": "Viejo titulo",
            "descripcion": "Vieja descripcion",
            "severidad": "leve",
            "justificacion_id": None,
        },
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_tipo_evento_by_id",
        lambda tipo_id: {"id": tipo_id, "activo": 1, "requiere_rango_fechas": 0},
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_tipos_evento",
        lambda include_inactive=False: [{"id": 1, "nombre": "Amonestacion"}],
    )

    captured = {}
    monkeypatch.setattr(
        legajos_routes,
        "update_evento",
        lambda evento_id, data: captured.update({"evento_id": evento_id, "data": dict(data)}),
    )

    resp = client.post(
        "/legajos/empleado/7/eventos/44/editar",
        data={
            "tipo_id": "1",
            "fecha_evento": "2026-03-06",
            "descripcion": "Descripcion editada",
            "titulo": "Titulo editado",
            "severidad": "media",
        },
        content_type="multipart/form-data",
        follow_redirects=False,
    )

    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/legajos/empleado/7")
    assert captured["evento_id"] == 44
    assert captured["data"]["descripcion"] == "Descripcion editada"
    assert captured["data"]["severidad"] == "media"


def test_legajos_anular_evento(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajos_routes, "log_audit", lambda *args, **kwargs: None)
    monkeypatch.setattr(
        legajos_routes,
        "get_empleado_by_id",
        lambda emp_id: {
            "id": emp_id,
            "empresa_id": 3,
            "legajo": "L-99",
            "dni": "30123456",
            "nombre": "Ana",
            "apellido": "Perez",
        },
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_evento_by_id",
        lambda evento_id: {
            "id": evento_id,
            "empleado_id": 7,
            "estado": "vigente",
        },
    )
    captured = {}
    monkeypatch.setattr(
        legajos_routes,
        "anular_evento",
        lambda evento_id, actor_id, motivo: captured.update(
            {"evento_id": evento_id, "actor_id": actor_id, "motivo": motivo}
        ),
    )

    resp = client.post(
        "/legajos/empleado/7/eventos/44/anular",
        data={"motivo_anulacion": "Documento invalido"},
        content_type="multipart/form-data",
        follow_redirects=False,
    )

    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/legajos/empleado/7")
    assert captured["evento_id"] == 44
    assert captured["actor_id"] == 11
    assert captured["motivo"] == "Documento invalido"


def test_legajos_eliminar_adjunto(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajos_routes, "log_audit", lambda *args, **kwargs: None)
    monkeypatch.setattr(
        legajos_routes,
        "get_empleado_by_id",
        lambda emp_id: {
            "id": emp_id,
            "empresa_id": 3,
            "legajo": "L-99",
            "dni": "30123456",
            "nombre": "Ana",
            "apellido": "Perez",
        },
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_evento_by_id",
        lambda evento_id: {
            "id": evento_id,
            "empleado_id": 7,
            "estado": "vigente",
        },
    )
    monkeypatch.setattr(
        legajos_routes,
        "get_adjunto_by_id",
        lambda adjunto_id: {
            "id": adjunto_id,
            "evento_id": 44,
            "estado": "activo",
        },
    )
    captured = {}
    monkeypatch.setattr(
        legajos_routes,
        "mark_deleted",
        lambda adjunto_id, actor_id: captured.update({"adjunto_id": adjunto_id, "actor_id": actor_id}),
    )

    resp = client.post(
        "/legajos/empleado/7/eventos/44/adjuntos/99/eliminar",
        data={},
        content_type="multipart/form-data",
        follow_redirects=False,
    )

    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/legajos/empleado/7")
    assert captured["adjunto_id"] == 99
    assert captured["actor_id"] == 11


def test_legajo_tipos_evento_listado(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        legajo_tipos_evento_routes,
        "get_tipos_evento_page",
        lambda **kwargs: (
            [
                {
                    "id": 1,
                    "codigo": "certificado_medico",
                    "nombre": "Certificado medico",
                    "requiere_rango_fechas": 0,
                    "permite_adjuntos": 1,
                    "habilitado_mobile": 1,
                    "activo": 1,
                }
            ],
            1,
        ),
    )

    resp = client.get("/legajos/tipos-evento/?activo=1")
    assert resp.status_code == 200
    assert b"Tipos de evento" in resp.data
    assert b"certificado_medico" in resp.data
    assert b"Habilitado" in resp.data


def test_legajo_tipos_evento_nuevo(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajo_tipos_evento_routes, "log_audit", lambda *args, **kwargs: None)
    monkeypatch.setattr(legajo_tipos_evento_routes, "get_tipo_evento_by_codigo", lambda codigo: None)

    captured = {}

    def _fake_create_tipo_evento(data):
        captured["data"] = dict(data)
        return 51

    monkeypatch.setattr(legajo_tipos_evento_routes, "create_tipo_evento", _fake_create_tipo_evento)

    resp = client.post(
        "/legajos/tipos-evento/nuevo",
        data={
            "codigo": " Certificado  Medico ",
            "nombre": "Certificado medico",
            "requiere_rango_fechas": "1",
            "permite_adjuntos": "1",
            "habilitado_mobile": "1",
            "activo": "1",
        },
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/legajos/tipos-evento/")
    assert captured["data"]["codigo"] == "certificado_medico"
    assert captured["data"]["requiere_rango_fechas"] is True
    assert captured["data"]["permite_adjuntos"] is True
    assert captured["data"]["habilitado_mobile"] is True
    assert captured["data"]["activo"] is True


def test_legajo_tipos_evento_editar(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajo_tipos_evento_routes, "log_audit", lambda *args, **kwargs: None)
    monkeypatch.setattr(
        legajo_tipos_evento_routes,
        "get_tipo_evento_by_id",
        lambda tipo_id: {
            "id": tipo_id,
            "codigo": "certificado_medico",
            "nombre": "Certificado medico",
            "requiere_rango_fechas": 0,
            "permite_adjuntos": 1,
            "habilitado_mobile": 0,
            "activo": 1,
        },
    )
    monkeypatch.setattr(
        legajo_tipos_evento_routes,
        "get_tipo_evento_by_codigo",
        lambda codigo: {"id": 9} if codigo == "certificado_medico" else None,
    )

    captured = {}
    monkeypatch.setattr(
        legajo_tipos_evento_routes,
        "update_tipo_evento",
        lambda tipo_id, data: captured.update({"tipo_id": tipo_id, "data": dict(data)}),
    )

    resp = client.post(
        "/legajos/tipos-evento/editar/9",
        data={
            "codigo": "certificado_medico",
            "nombre": "Certificado actualizado",
            "permite_adjuntos": "1",
            "habilitado_mobile": "1",
            "activo": "1",
        },
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/legajos/tipos-evento/")
    assert captured["tipo_id"] == 9
    assert captured["data"]["nombre"] == "Certificado actualizado"
    assert captured["data"]["requiere_rango_fechas"] is False
    assert captured["data"]["habilitado_mobile"] is True


def test_legajo_tipos_evento_activar_desactivar(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajo_tipos_evento_routes, "log_audit", lambda *args, **kwargs: None)
    monkeypatch.setattr(legajo_tipos_evento_routes, "get_tipo_evento_by_id", lambda tipo_id: {"id": tipo_id})
    monkeypatch.setattr(legajo_tipos_evento_routes, "count_eventos_vigentes_by_tipo", lambda tipo_id: 0)

    captured = []
    monkeypatch.setattr(
        legajo_tipos_evento_routes,
        "set_tipo_evento_activo",
        lambda tipo_id, activo: captured.append((tipo_id, activo)),
    )

    resp_1 = client.get("/legajos/tipos-evento/desactivar/9", follow_redirects=False)
    resp_2 = client.get("/legajos/tipos-evento/activar/9", follow_redirects=False)

    assert resp_1.status_code == 302
    assert resp_2.status_code == 302
    assert captured == [(9, 0), (9, 1)]


def test_legajos_permisos_mobile_listado(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajos_routes, "get_empleados", lambda include_inactive=False: [])
    monkeypatch.setattr(
        legajos_routes,
        "get_mobile_legajo_permisos_page",
        lambda page, per_page, *, search=None, activo=None: (
            [
                {
                    "id": 4,
                    "apellido": "Perez",
                    "nombre": "Ana",
                    "legajo": "L-1",
                    "dni": "301",
                    "empresa_nombre": "Empresa A",
                    "sucursal_nombre": "Centro",
                    "sector_nombre": "Operaciones",
                    "alcance": "sector",
                    "activo": 1,
                }
            ],
            1,
        ),
    )

    resp = client.get("/legajos/permisos-mobile")

    assert resp.status_code == 200
    assert b"Permisos mobile" in resp.data
    assert b"Perez" in resp.data
    assert b"Sector" in resp.data


def test_legajos_permisos_mobile_asigna_permiso(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajos_routes, "log_audit", lambda *args, **kwargs: None)
    captured = {}
    monkeypatch.setattr(
        legajos_routes,
        "upsert_mobile_legajo_permiso",
        lambda empleado_id, *, alcance, activo=1: captured.update(
            {"empleado_id": empleado_id, "alcance": alcance, "activo": activo}
        ),
    )

    resp = client.post(
        "/legajos/permisos-mobile",
        data={"empleado_id": "7", "alcance": "sucursal"},
        follow_redirects=False,
    )

    assert resp.status_code == 302
    assert captured == {"empleado_id": 7, "alcance": "sucursal", "activo": 1}


def test_legajos_permisos_mobile_desactiva(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajos_routes, "log_audit", lambda *args, **kwargs: None)
    captured = {}
    monkeypatch.setattr(
        legajos_routes,
        "set_mobile_legajo_permiso_activo",
        lambda permiso_id, activo: captured.update({"permiso_id": permiso_id, "activo": activo}) or True,
    )

    resp = client.post("/legajos/permisos-mobile/4/desactivar", follow_redirects=False)

    assert resp.status_code == 302
    assert captured == {"permiso_id": 4, "activo": 0}


def test_legajo_tipos_evento_no_desactiva_con_eventos_vigentes(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(legajo_tipos_evento_routes, "get_tipo_evento_by_id", lambda tipo_id: {"id": tipo_id})
    monkeypatch.setattr(legajo_tipos_evento_routes, "count_eventos_vigentes_by_tipo", lambda tipo_id: 3)

    called = {"set": False}
    monkeypatch.setattr(
        legajo_tipos_evento_routes,
        "set_tipo_evento_activo",
        lambda tipo_id, activo: called.update({"set": True}),
    )

    resp = client.get("/legajos/tipos-evento/desactivar/9", follow_redirects=False)
    assert resp.status_code == 400
    assert b"eventos vigentes asociados" in resp.data
    assert called["set"] is False
