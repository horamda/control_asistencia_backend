import io
import datetime

from openpyxl import load_workbook

import app as app_module
import web.auth.decorators as auth_decorators
import web.asistencias.asistencias_routes as asistencias_routes


def _build_client(monkeypatch):
    monkeypatch.setattr(app_module, "init_db", lambda: None)
    app = app_module.create_app()
    app.config["TESTING"] = True
    app.config["WTF_CSRF_ENABLED"] = False
    return app.test_client()


def _login_session(client):
    with client.session_transaction() as sess:
        sess["user_id"] = 99


def test_asistencias_get_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(asistencias_routes, "get_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(asistencias_routes, "get_empleados", lambda include_inactive=True: [])
    monkeypatch.setattr(asistencias_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(asistencias_routes, "get_sectores", lambda include_inactive=True: [])

    resp = client.get("/asistencias/")
    assert resp.status_code == 200
    assert b"Generar ausentes por rango" in resp.data


def test_asistencias_get_muestra_error(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(asistencias_routes, "get_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(asistencias_routes, "get_empleados", lambda include_inactive=True: [])
    monkeypatch.setattr(asistencias_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(asistencias_routes, "get_sectores", lambda include_inactive=True: [])

    resp = client.get("/asistencias/?error=Fecha+invalida")
    assert resp.status_code == 200
    assert b"Fecha invalida" in resp.data


def test_generar_ausentes_dia(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    captured = {}

    def _fake_generar_ausentes(fecha):
        captured["fecha"] = fecha
        return 0, []

    monkeypatch.setattr(asistencias_routes, "generar_ausentes", _fake_generar_ausentes)
    monkeypatch.setattr(asistencias_routes, "generar_ausentes_rango", lambda *args, **kwargs: (0, []))

    fecha = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()
    resp = client.post(
        "/asistencias/generar-ausentes",
        data={"modo": "dia", "fecha": fecha},
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert f"fecha_desde={fecha}" in resp.headers["Location"]
    assert captured["fecha"] == fecha


def test_reportes_mensuales_renderiza_solapas(monkeypatch):
    monkeypatch.setattr(auth_decorators, "has_role", lambda user_id, role: True)
    monkeypatch.setattr(asistencias_routes, "has_role", lambda user_id, role: True)
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(asistencias_routes, "get_empresas", lambda include_inactive=True: [{"id": 1, "razon_social": "Empresa"}])
    monkeypatch.setattr(asistencias_routes, "get_sucursales", lambda include_inactive=True: [{"id": 2, "empresa_id": 1, "nombre": "Suc. Dolores"}])
    monkeypatch.setattr(asistencias_routes, "get_sectores", lambda include_inactive=True: [{"id": 5, "empresa_id": 1, "nombre": "Operaciones"}])
    monkeypatch.setattr(asistencias_routes, "get_dias_no_laborables", lambda **kwargs: set())
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True, sucursal_id=None, sector_id=None: [
            {"id": 10, "empresa_id": 1, "sucursal_id": 2, "apellido": "Aguirre", "nombre": "Leandro", "sector_nombre": "Operaciones", "activo": 1}
        ],
    )
    monkeypatch.setattr(asistencias_routes, "get_marcas_admin_export", lambda **kwargs: [])
    monkeypatch.setattr(asistencias_routes, "get_justificaciones_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(asistencias_routes, "get_vacaciones_aprobadas_export", lambda **kwargs: [])

    resp = client.get("/asistencias/reportes?mes=2026-07&sucursal_id=2")

    assert resp.status_code == 200
    assert b"Julio 2026" in resp.data
    assert b"Resumen" in resp.data
    assert b"Ausencias" in resp.data
    assert b"An" in resp.data
    assert b"Jornada" in resp.data
    assert b"Aguirre Leandro" in resp.data


def test_reportes_mensuales_usa_dias_no_laborables_guardados(monkeypatch):
    monkeypatch.setattr(auth_decorators, "has_role", lambda user_id, role: True)
    monkeypatch.setattr(asistencias_routes, "has_role", lambda user_id, role: True)
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(asistencias_routes, "get_empresas", lambda include_inactive=True: [{"id": 1, "razon_social": "Empresa"}])
    monkeypatch.setattr(asistencias_routes, "get_sucursales", lambda include_inactive=True: [{"id": 2, "empresa_id": 1, "nombre": "Suc. Dolores"}])
    monkeypatch.setattr(asistencias_routes, "get_sectores", lambda include_inactive=True: [])
    monkeypatch.setattr(
        asistencias_routes,
        "get_dias_no_laborables",
        lambda **kwargs: {"2026-07-13"},
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True, sucursal_id=None, sector_id=None: [
            {"id": 10, "empresa_id": 1, "sucursal_id": 2, "apellido": "Aguirre", "nombre": "Leandro", "sector_nombre": "Operaciones", "activo": 1}
        ],
    )
    monkeypatch.setattr(asistencias_routes, "get_marcas_admin_export", lambda **kwargs: [])
    monkeypatch.setattr(asistencias_routes, "get_justificaciones_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(asistencias_routes, "get_vacaciones_aprobadas_export", lambda **kwargs: [])

    resp = client.get("/asistencias/reportes?mes=2026-07&empresa_id=1&sucursal_id=2")
    html = resp.get_data(as_text=True)

    assert resp.status_code == 200
    assert 'name="nl" value="5,12,13,19,26"' in html
    assert 'data-day="13">13</button>' in html
    assert 'non-laborable" data-day="5">5</button>' in html
    assert 'non-laborable" data-day="12">12</button>' in html
    assert 'non-laborable" data-day="19">19</button>' in html
    assert 'non-laborable" data-day="26">26</button>' in html


def test_reportes_mensuales_filtra_por_sector(monkeypatch):
    monkeypatch.setattr(auth_decorators, "has_role", lambda user_id, role: True)
    monkeypatch.setattr(asistencias_routes, "has_role", lambda user_id, role: True)
    client = _build_client(monkeypatch)
    _login_session(client)
    captured = {}
    monkeypatch.setattr(asistencias_routes, "get_empresas", lambda include_inactive=True: [{"id": 1, "razon_social": "Empresa"}])
    monkeypatch.setattr(asistencias_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(asistencias_routes, "get_sectores", lambda include_inactive=True: [{"id": 5, "empresa_id": 1, "nombre": "Operaciones"}])
    monkeypatch.setattr(asistencias_routes, "get_dias_no_laborables", lambda **kwargs: set())

    def _fake_get_empleados(include_inactive=True, sucursal_id=None, sector_id=None):
        captured["sector_id"] = sector_id
        return []

    monkeypatch.setattr(asistencias_routes, "get_empleados", _fake_get_empleados)
    monkeypatch.setattr(asistencias_routes, "get_marcas_admin_export", lambda **kwargs: [])
    monkeypatch.setattr(asistencias_routes, "get_justificaciones_page", lambda *args, **kwargs: ([], 0))
    monkeypatch.setattr(asistencias_routes, "get_vacaciones_aprobadas_export", lambda **kwargs: [])

    resp = client.get("/asistencias/reportes?mes=2026-07&empresa_id=1&sector_id=5")

    assert resp.status_code == 200
    assert captured["sector_id"] == 5


def test_reportes_dias_no_laborables_guardar(monkeypatch):
    monkeypatch.setattr(auth_decorators, "has_role", lambda user_id, role: True)
    client = _build_client(monkeypatch)
    _login_session(client)
    captured = {}

    def _fake_replace(**kwargs):
        captured.update(kwargs)
        return len(kwargs["dates"])

    monkeypatch.setattr(asistencias_routes, "replace_dias_no_laborables", _fake_replace)

    resp = client.post(
        "/asistencias/reportes/dias-no-laborables",
        data={
            "mes": "2026-07",
            "empresa_id": "1",
            "sucursal_id": "2",
            "sector_id": "5",
            "tab": "resumen",
            "nl": "13,20,99,abc",
        },
        follow_redirects=False,
    )

    assert resp.status_code == 302
    assert captured["year"] == 2026
    assert captured["month"] == 7
    assert captured["empresa_id"] == 1
    assert captured["sucursal_id"] == 2
    assert captured["sector_id"] == 5
    assert captured["actor_id"] == 99
    assert captured["dates"] == {"2026-07-13", "2026-07-20"}
    assert "sector_id=5" in resp.headers["Location"]


def test_generar_ausentes_rango(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    captured = {}

    def _fake_generar_ausentes_rango(fecha_desde, fecha_hasta):
        captured["desde"] = fecha_desde
        captured["hasta"] = fecha_hasta
        return 0, []

    monkeypatch.setattr(asistencias_routes, "generar_ausentes", lambda *args, **kwargs: (0, []))
    monkeypatch.setattr(asistencias_routes, "generar_ausentes_rango", _fake_generar_ausentes_rango)

    fecha_hasta = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()
    fecha_desde = (datetime.date.today() - datetime.timedelta(days=10)).isoformat()
    resp = client.post(
        "/asistencias/generar-ausentes",
        data={"modo": "rango", "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta},
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert f"fecha_desde={fecha_desde}" in resp.headers["Location"]
    assert f"fecha_hasta={fecha_hasta}" in resp.headers["Location"]
    assert captured["desde"] == fecha_desde
    assert captured["hasta"] == fecha_hasta


def test_generar_ausentes_rango_propagates_errors(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(asistencias_routes, "generar_ausentes", lambda *args, **kwargs: (0, []))
    monkeypatch.setattr(
        asistencias_routes,
        "generar_ausentes_rango",
        lambda fecha_desde, fecha_hasta: (0, ["fecha_desde no puede ser mayor a fecha_hasta."]),
    )

    fecha_hasta = (datetime.date.today() - datetime.timedelta(days=5)).isoformat()
    fecha_desde = (datetime.date.today() - datetime.timedelta(days=2)).isoformat()
    resp = client.post(
        "/asistencias/generar-ausentes",
        data={"modo": "rango", "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta},
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert "error=fecha_desde+no+puede+ser+mayor+a+fecha_hasta." in resp.headers["Location"]


def test_generar_ausentes_rango_fecha_hasta_futura(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    called = {"rango": False}

    def _fake_generar_ausentes_rango(fecha_desde, fecha_hasta):
        called["rango"] = True
        return 0, []

    monkeypatch.setattr(asistencias_routes, "generar_ausentes", lambda *args, **kwargs: (0, []))
    monkeypatch.setattr(asistencias_routes, "generar_ausentes_rango", _fake_generar_ausentes_rango)

    fecha_desde = (datetime.date.today() - datetime.timedelta(days=2)).isoformat()
    fecha_hasta = (datetime.date.today() + datetime.timedelta(days=1)).isoformat()
    resp = client.post(
        "/asistencias/generar-ausentes",
        data={"modo": "rango", "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta},
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert "error=" in resp.headers["Location"]
    assert "fecha_hasta+no+puede+ser+mayor+a+hoy" in resp.headers["Location"]
    assert called["rango"] is False


def test_sync_simple_marcas_for_asistencia_crea_par_basico(monkeypatch):
    monkeypatch.setattr(
        asistencias_routes,
        "get_by_id",
        lambda asistencia_id: {
            "id": asistencia_id,
            "empresa_id": 1,
            "empleado_id": 100,
            "fecha": "2026-03-10",
            "hora_entrada": "08:00:00",
            "hora_salida": "12:00:00",
            "metodo_entrada": "manual",
            "metodo_salida": "manual",
            "lat_entrada": None,
            "lon_entrada": None,
            "lat_salida": None,
            "lon_salida": None,
            "foto_entrada": None,
            "foto_salida": None,
            "gps_ok_entrada": None,
            "gps_ok_salida": None,
            "gps_distancia_entrada_m": None,
            "gps_distancia_salida_m": None,
            "gps_tolerancia_entrada_m": None,
            "gps_tolerancia_salida_m": None,
            "gps_ref_lat_entrada": None,
            "gps_ref_lon_entrada": None,
            "gps_ref_lat_salida": None,
            "gps_ref_lon_salida": None,
            "estado": "ok",
            "observaciones": "manual",
        },
    )
    monkeypatch.setattr(asistencias_routes, "get_marcas_by_asistencia", lambda asistencia_id: [])
    deleted = {"count": 0}
    created = {"rows": []}

    monkeypatch.setattr(
        asistencias_routes,
        "delete_marca_by_id",
        lambda marca_id: deleted.__setitem__("count", deleted["count"] + 1) or True,
    )

    def _fake_create_marca(**kwargs):
        created["rows"].append(kwargs)
        return len(created["rows"])

    monkeypatch.setattr(asistencias_routes, "create_marca", _fake_create_marca)

    result = asistencias_routes._sync_simple_marcas_for_asistencia(77)
    assert result["synced"] is True
    assert result["created"] == 2
    assert result["deleted"] == 0
    assert len(created["rows"]) == 2
    assert created["rows"][0]["accion"] == "ingreso"
    assert created["rows"][1]["accion"] == "egreso"
    assert deleted["count"] == 0


def test_sync_simple_marcas_for_asistencia_saltea_si_hay_multiples(monkeypatch):
    monkeypatch.setattr(
        asistencias_routes,
        "get_by_id",
        lambda asistencia_id: {
            "id": asistencia_id,
            "empresa_id": 1,
            "empleado_id": 100,
            "fecha": "2026-03-10",
            "hora_entrada": "08:00:00",
            "hora_salida": "12:00:00",
        },
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_marcas_by_asistencia",
        lambda asistencia_id: [
            {"id": 1, "accion": "ingreso", "tipo_marca": "jornada"},
            {"id": 2, "accion": "ingreso", "tipo_marca": "jornada"},
        ],
    )

    result = asistencias_routes._sync_simple_marcas_for_asistencia(77)
    assert result["synced"] is False
    assert result["reason"] == "multiple_marcas"


def test_asistencias_nuevo_dispara_sync_automatico(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True: [{"id": 100, "apellido": "Perez", "nombre": "Ana", "dni": "30111222"}],
    )
    monkeypatch.setattr(asistencias_routes, "_validate", lambda form: [])
    monkeypatch.setattr(asistencias_routes, "validar_asistencia", lambda *args, **kwargs: ([], "ok"))
    monkeypatch.setattr(asistencias_routes, "create", lambda data: 321)
    monkeypatch.setattr(asistencias_routes, "log_audit", lambda *args, **kwargs: True)
    captured = {}
    monkeypatch.setattr(
        asistencias_routes,
        "_sync_simple_marcas_for_asistencia",
        lambda asistencia_id: captured.__setitem__("id", asistencia_id) or {"synced": True},
    )

    resp = client.post(
        "/asistencias/nuevo",
        data={
            "empleado_id": "100",
            "fecha": "2026-03-10",
            "hora_entrada": "08:00",
            "hora_salida": "12:00",
            "metodo_entrada": "manual",
            "metodo_salida": "manual",
        },
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert "/asistencias/" in resp.headers["Location"]
    assert captured["id"] == 321


def test_asistencias_editar_dispara_sync_automatico(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        asistencias_routes,
        "get_by_id",
        lambda asistencia_id: {"id": asistencia_id, "empleado_id": 100, "fecha": "2026-03-10"},
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True: [{"id": 100, "apellido": "Perez", "nombre": "Ana", "dni": "30111222"}],
    )
    monkeypatch.setattr(asistencias_routes, "_validate", lambda form: [])
    monkeypatch.setattr(asistencias_routes, "validar_asistencia", lambda *args, **kwargs: ([], "ok"))
    monkeypatch.setattr(asistencias_routes, "update", lambda asistencia_id, data: True)
    monkeypatch.setattr(asistencias_routes, "log_audit", lambda *args, **kwargs: True)
    captured = {}
    monkeypatch.setattr(
        asistencias_routes,
        "_sync_simple_marcas_for_asistencia",
        lambda asistencia_id: captured.__setitem__("id", asistencia_id) or {"synced": True},
    )

    resp = client.post(
        "/asistencias/editar/321",
        data={
            "empleado_id": "100",
            "fecha": "2026-03-10",
            "hora_entrada": "08:00",
            "hora_salida": "12:00",
            "metodo_entrada": "manual",
            "metodo_salida": "manual",
        },
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert "/asistencias/" in resp.headers["Location"]
    assert captured["id"] == 321


def test_historial_marcas_get_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(asistencias_routes, "get_marcas_admin_page", lambda **kwargs: ([], 0))
    monkeypatch.setattr(asistencias_routes, "get_empleados", lambda include_inactive=True: [])
    monkeypatch.setattr(asistencias_routes, "get_empresas", lambda include_inactive=True: [])

    resp = client.get("/asistencias/marcas")
    assert resp.status_code == 200
    assert b"Historial de marcas" in resp.data
    assert b"Reporte CSV" in resp.data


def test_historial_marcas_csv_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        asistencias_routes,
        "get_marcas_admin_export",
        lambda **kwargs: [
            {
                "id": 1,
                "empresa_nombre": "Empresa Test",
                "apellido": "Perez",
                "nombre": "Ana",
                "dni": "30111222",
                "fecha": "2026-02-21",
                "hora": "08:00:00",
                "accion": "ingreso",
                "tipo_marca": "jornada",
                "metodo": "qr",
                "gps_ok": 1,
                "gps_distancia_m": 3.2,
                "gps_tolerancia_m": 30.0,
                "lat": -34.6,
                "lon": -58.4,
                "estado": "ok",
                "observaciones": "",
                "fecha_creacion": "2026-02-21 08:00:01",
            }
        ],
    )

    resp = client.get("/asistencias/marcas.csv")
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["Content-Type"]
    assert "historial_marcas_" in resp.headers["Content-Disposition"]
    assert b"Empresa Test" in resp.data


def test_historial_marcas_reporte_csv_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    captured = {}

    rows = [
        {
            "id": 2,
            "empleado_id": 20,
            "fecha": datetime.date(2026, 6, 11),
            "hora": datetime.time(16, 54),
            "accion": "egreso",
            "legajo": "803",
            "apellido": "Gonzalez",
            "nombre": "Lucas",
            "sucursal_nombre": "Porton Lateral",
            "sector_nombre": "Almacen",
            "dni": "803",
        },
        {
            "id": 1,
            "empleado_id": 10,
            "fecha": datetime.date(2026, 6, 11),
            "hora": datetime.time(13, 50),
            "accion": "ingreso",
            "legajo": "58",
            "apellido": "Pereyra",
            "nombre": "Gabriel",
            "sucursal_nombre": "Porton Lateral",
            "sector_nombre": "Reparto",
            "dni": "58",
        },
    ]

    def _fake_export(**kwargs):
        captured["kwargs"] = kwargs
        if kwargs.get("order_asc"):
            return list(reversed(rows))
        return list(rows)

    monkeypatch.setattr(asistencias_routes, "get_marcas_admin_export", _fake_export)

    resp = client.get("/asistencias/marcas/reporte.csv?empresa_id=1&fecha_desde=2026-06-11&fecha_hasta=2026-06-11")
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["Content-Type"]
    assert "reporte_asistencia_" in resp.headers["Content-Disposition"]
    assert captured["kwargs"]["order_asc"] is True

    lines = resp.data.decode("utf-8-sig").splitlines()
    assert lines[0] == "MES,FECHA,HORA,PUERTA,TIPO MOV,CODIGO,NOMBRE,SECTOR"
    assert lines[1] == "6,11/6/2026,13:50,Porton Lateral,Entrada,58,PEREYRA GABRIEL,Reparto"
    assert lines[2] == "6,11/6/2026,16:54,Porton Lateral,Salida,803,GONZALEZ LUCAS,Almacen"


def test_historial_marcas_xlsx_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        asistencias_routes,
        "get_empresas",
        lambda include_inactive=True: [{"id": 1, "razon_social": "Empresa Test"}],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True: [{"id": 2, "apellido": "Perez", "nombre": "Ana"}],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_marcas_admin_export",
        lambda **kwargs: [
            {
                "id": 1,
                "empresa_nombre": "Empresa Test",
                "apellido": "Perez",
                "nombre": "Ana",
                "dni": "30111222",
                "fecha": "2026-02-21",
                "hora": "08:00:00",
                "accion": "ingreso",
                "tipo_marca": "jornada",
                "metodo": "qr",
                "gps_ok": 1,
                "gps_distancia_m": 3.2,
                "gps_tolerancia_m": 30.0,
                "lat": -34.6,
                "lon": -58.4,
                "estado": "ok",
                "observaciones": "",
                "fecha_creacion": "2026-02-21 08:00:01",
            }
        ],
    )

    resp = client.get("/asistencias/marcas.xlsx?empresa_id=1&empleado_id=2")
    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp.headers["Content-Type"]
    assert "historial_marcas_" in resp.headers["Content-Disposition"]

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb["Historial"]
    assert ws["A1"].value == "Historial de marcas"
    assert ws["A5"].value == "Empresa"
    assert ws["B5"].value == "Empresa Test"
    assert ws["A16"].value == "ID"
    assert ws["B17"].value == "Empresa Test"


def test_historial_marcas_backfill_post_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(asistencias_routes, "backfill_marcas", lambda: (4, 2))

    resp = client.post(
        "/asistencias/marcas/backfill",
        data={
            "page": "1",
            "per": "20",
            "empresa_id": "",
            "empleado_id": "",
            "fecha_desde": "",
            "fecha_hasta": "",
            "tipo_marca": "",
            "accion": "",
            "metodo": "",
            "gps_ok": "",
            "q": "",
        },
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert "/asistencias/marcas" in resp.headers["Location"]
    assert "backfill_ingresos=4" in resp.headers["Location"]
    assert "backfill_egresos=2" in resp.headers["Location"]


def test_planilla_diaria_get_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        asistencias_routes,
        "get_empresas",
        lambda include_inactive=True: [{"id": 1, "razon_social": "Del Palacio S.A"}],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_sucursales",
        lambda include_inactive=True: [{"id": 10, "empresa_id": 1, "nombre": "Casa Central"}],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True: [
            {"id": 100, "empresa_id": 1, "sucursal_id": 10, "apellido": "Persona", "nombre": "Uno", "dni": "123"}
        ],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_marcas_admin_export",
        lambda **kwargs: [
            {
                "id": 1,
                "empleado_id": 100,
                "asistencia_id": 77,
                "hora": "07:00:00",
                "accion": "ingreso",
            },
            {
                "id": 2,
                "empleado_id": 100,
                "asistencia_id": 77,
                "hora": "12:00:00",
                "accion": "egreso",
            },
        ],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_configuracion_empresa_by_id",
        lambda empresa_id: {"intervalo_minimo_fichadas_minutos": 60},
    )
    monkeypatch.setattr(asistencias_routes, "get_page", lambda *args, **kwargs: ([], 0))

    resp = client.get("/asistencias/planilla?empresa_id=1&sucursal_id=10&fecha=2026-03-10")
    assert resp.status_code == 200
    assert b"Planilla diaria de fichadas" in resp.data
    assert b"Del Palacio S.A" in resp.data
    assert b"Casa Central" in resp.data
    assert b"07:00" in resp.data
    assert b"12:00" in resp.data
    html = resp.get_data(as_text=True)
    assert "/asistencias/planilla/marca/editar/1" in html
    assert "/asistencias/planilla/marca/eliminar/1" in html


def test_planilla_diaria_detecta_intervalo_corto(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(asistencias_routes, "get_empresas", lambda include_inactive=True: [{"id": 1, "razon_social": "X"}])
    monkeypatch.setattr(asistencias_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True: [
            {"id": 101, "empresa_id": 1, "sucursal_id": None, "apellido": "Persona", "nombre": "Dos", "dni": "222"}
        ],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_marcas_admin_export",
        lambda **kwargs: [
            {"id": 1, "empleado_id": 101, "asistencia_id": 88, "hora": "07:00:00", "accion": "ingreso"},
            {"id": 2, "empleado_id": 101, "asistencia_id": 88, "hora": "07:10:00", "accion": "egreso"},
        ],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_configuracion_empresa_by_id",
        lambda empresa_id: {"intervalo_minimo_fichadas_minutos": 60},
    )
    monkeypatch.setattr(asistencias_routes, "get_page", lambda *args, **kwargs: ([], 0))

    resp = client.get("/asistencias/planilla?empresa_id=1&fecha=2026-03-10")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "Intervalo corto" in html


def test_planilla_diaria_normaliza_hora_hhmm(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(asistencias_routes, "get_empresas", lambda include_inactive=True: [{"id": 1, "razon_social": "X"}])
    monkeypatch.setattr(asistencias_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True: [
            {"id": 101, "empresa_id": 1, "sucursal_id": None, "apellido": "Persona", "nombre": "Dos", "dni": "222"}
        ],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_marcas_admin_export",
        lambda **kwargs: [
            {"id": 1, "empleado_id": 101, "asistencia_id": 88, "hora": "8:01:00", "accion": "ingreso"},
        ],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_configuracion_empresa_by_id",
        lambda empresa_id: {"intervalo_minimo_fichadas_minutos": 60},
    )
    monkeypatch.setattr(asistencias_routes, "get_page", lambda *args, **kwargs: ([], 0))

    resp = client.get("/asistencias/planilla?empresa_id=1&fecha=2026-03-10")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "08:01" in html


def test_planilla_diaria_export_excel_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        asistencias_routes,
        "get_empresas",
        lambda include_inactive=True: [{"id": 1, "razon_social": "Del Palacio S.A"}],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_sucursales",
        lambda include_inactive=True: [{"id": 10, "empresa_id": 1, "nombre": "Casa Central"}],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True: [
            {"id": 100, "empresa_id": 1, "sucursal_id": 10, "apellido": "Persona", "nombre": "Uno", "dni": "123"}
        ],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_marcas_admin_export",
        lambda **kwargs: [
            {"id": 1, "empleado_id": 100, "asistencia_id": 77, "hora": "07:00:00", "accion": "ingreso"},
            {"id": 2, "empleado_id": 100, "asistencia_id": 77, "hora": "12:00:00", "accion": "egreso"},
        ],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_configuracion_empresa_by_id",
        lambda empresa_id: {"intervalo_minimo_fichadas_minutos": 60},
    )
    monkeypatch.setattr(asistencias_routes, "get_page", lambda *args, **kwargs: ([], 0))

    resp = client.get("/asistencias/planilla.xls?empresa_id=1&sucursal_id=10&fecha=2026-03-10")
    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers["Content-Type"]
    assert "planilla_fichadas_2026-03-10.xlsx" in resp.headers["Content-Disposition"]

    wb = load_workbook(io.BytesIO(resp.data), data_only=True)
    assert wb.sheetnames == ["Planilla"]

    ws = wb["Planilla"]
    assert ws["A1"].value == "Planilla diaria de fichadas"
    assert ws["A5"].value == "Empleado"
    assert ws["B5"].value == "DNI"
    assert ws["C5"].value == "Ingreso 1"
    assert ws["D5"].value == "Egreso 1"
    assert ws["A6"].value == "Persona Uno"
    assert ws["B6"].value == "123"
    assert ws["C6"].value == "07:00"
    assert ws["D6"].value == "12:00"


def test_planilla_diaria_export_pdf_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(asistencias_routes, "get_empresas", lambda include_inactive=True: [{"id": 1, "razon_social": "X"}])
    monkeypatch.setattr(asistencias_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True: [
            {"id": 101, "empresa_id": 1, "sucursal_id": None, "apellido": "Persona", "nombre": "Dos", "dni": "222"}
        ],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_marcas_admin_export",
        lambda **kwargs: [
            {"id": 1, "empleado_id": 101, "asistencia_id": 88, "hora": "07:00:00", "accion": "ingreso"},
            {"id": 2, "empleado_id": 101, "asistencia_id": 88, "hora": "12:00:00", "accion": "egreso"},
        ],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_configuracion_empresa_by_id",
        lambda empresa_id: {"intervalo_minimo_fichadas_minutos": 60},
    )
    monkeypatch.setattr(asistencias_routes, "get_page", lambda *args, **kwargs: ([], 0))

    resp = client.get("/asistencias/planilla.pdf?empresa_id=1&fecha=2026-03-10&auto_print=1")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "text/html" in resp.headers["Content-Type"]
    assert "window.print" in html
    assert "Planilla diaria de fichadas" in html


def test_planilla_marca_editar_get_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        asistencias_routes,
        "get_marca_by_id",
        lambda marca_id: {"id": marca_id, "asistencia_id": 77, "fecha": "2026-03-10", "hora": "08:00:00", "accion": "ingreso"},
    )

    resp = client.get("/asistencias/planilla/marca/editar/1?empresa_id=1&sucursal_id=10&fecha=2026-03-10")
    assert resp.status_code == 200
    assert b"Editar marca" in resp.data
    assert b"08:00" in resp.data


def test_planilla_diaria_muestra_asistencia_manual_sin_marcas(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        asistencias_routes,
        "get_empresas",
        lambda include_inactive=True: [{"id": 1, "razon_social": "Del Palacio S.A"}],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_sucursales",
        lambda include_inactive=True: [{"id": 10, "empresa_id": 1, "nombre": "Casa Central"}],
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_empleados",
        lambda include_inactive=True: [
            {"id": 200, "empresa_id": 1, "sucursal_id": 10, "apellido": "Manual", "nombre": "SinMarca", "dni": "456"}
        ],
    )
    monkeypatch.setattr(asistencias_routes, "get_marcas_admin_export", lambda **kwargs: [])
    monkeypatch.setattr(
        asistencias_routes,
        "get_page",
        lambda *args, **kwargs: (
            [
                {
                    "id": 901,
                    "empleado_id": 200,
                    "fecha": "2026-03-10",
                    "hora_entrada": "08:00:00",
                    "hora_salida": "12:00:00",
                    "gps_ok_entrada": 1,
                    "gps_ok_salida": 1,
                }
            ],
            1,
        ),
    )
    monkeypatch.setattr(
        asistencias_routes,
        "get_configuracion_empresa_by_id",
        lambda empresa_id: {"intervalo_minimo_fichadas_minutos": 60},
    )

    resp = client.get("/asistencias/planilla?empresa_id=1&sucursal_id=10&fecha=2026-03-10")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "Manual SinMarca" in html
    assert "08:00" in html
    assert "12:00" in html


def test_planilla_marca_eliminar_post_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        asistencias_routes,
        "get_marca_by_id",
        lambda marca_id: {"id": marca_id, "asistencia_id": 77, "fecha": "2026-03-10", "hora": "08:00:00", "accion": "ingreso"},
    )
    deleted = {"ok": False}
    synced = {"ok": False}
    monkeypatch.setattr(asistencias_routes, "delete_marca_by_id", lambda marca_id: deleted.__setitem__("ok", True) or True)
    monkeypatch.setattr(
        asistencias_routes, "sync_from_asistencia_marcas", lambda asistencia_id: synced.__setitem__("ok", True) or True
    )
    monkeypatch.setattr(asistencias_routes, "log_audit", lambda *args, **kwargs: True)

    resp = client.post(
        "/asistencias/planilla/marca/eliminar/1",
        data={"empresa_id": "1", "sucursal_id": "10", "fecha": "2026-03-10"},
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert "/asistencias/planilla" in resp.headers["Location"]
    assert "msg=Marca+%231+eliminada." in resp.headers["Location"]
    assert deleted["ok"] is True
    assert synced["ok"] is True


def test_planilla_marca_agregar_post_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)
    monkeypatch.setattr(
        asistencias_routes,
        "get_by_id",
        lambda asistencia_id: {
            "id": asistencia_id,
            "empresa_id": 1,
            "empleado_id": 101,
            "fecha": "2026-03-10",
            "lat_salida": None,
            "lon_salida": None,
            "foto_salida": None,
            "metodo_salida": "manual",
            "gps_ok_salida": None,
            "gps_distancia_salida_m": None,
            "gps_tolerancia_salida_m": None,
            "gps_ref_lat_salida": None,
            "gps_ref_lon_salida": None,
            "estado": "ok",
        },
    )
    monkeypatch.setattr(asistencias_routes, "create_marca", lambda **kwargs: 999)
    monkeypatch.setattr(asistencias_routes, "sync_from_asistencia_marcas", lambda asistencia_id: True)
    monkeypatch.setattr(asistencias_routes, "log_audit", lambda *args, **kwargs: True)

    resp = client.post(
        "/asistencias/planilla/marca/agregar",
        data={
            "asistencia_id": "77",
            "accion": "egreso",
            "hora": "12:00",
            "empresa_id": "1",
            "sucursal_id": "10",
            "fecha": "2026-03-10",
        },
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert "/asistencias/planilla" in resp.headers["Location"]
    assert "msg=Marca+%23999+agregada." in resp.headers["Location"]
