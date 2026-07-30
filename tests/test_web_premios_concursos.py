import io

from openpyxl import load_workbook

import app as app_module
import web.auth.decorators as auth_decorators
import web.premios_concursos.premios_concursos_routes as premios_routes


def _build_client(monkeypatch):
    monkeypatch.setattr(app_module, "init_db", lambda: None)
    app = app_module.create_app()
    app.config["TESTING"] = True
    app.config["WTF_CSRF_ENABLED"] = False
    return app.test_client()


def _login(client):
    with client.session_transaction() as sess:
        sess["user_id"] = 99
        sess["role"] = "admin"
        sess["nombre"] = "Test"


def _build_authed_client(monkeypatch):
    monkeypatch.setattr(auth_decorators, "has_role", lambda user_id, role: True)
    client = _build_client(monkeypatch)
    _login(client)
    return client


def _stub_empresas():
    return [{"id": 1, "razon_social": "Acme"}]


def _stub_sectores(*args, **kwargs):
    return [{"id": 3, "nombre": "Logistica"}]


def _stub_empleados(*args, **kwargs):
    return [{"id": 10, "apellido": "Lopez", "nombre": "Ana", "legajo": "L001", "dni": "123"}]


def _stub_concursos(*args, **kwargs):
    return [
        {
            "id": 9,
            "empresa_id": 1,
            "sector_id": None,
            "codigo": "SEGURIDAD",
            "nombre": "Premio de seguridad",
            "descripcion": None,
            "alcance": "global",
            "activo": 1,
            "empresa_nombre": "Acme",
            "sector_nombre": None,
        }
    ]


def test_premios_concursos_listado_ok(monkeypatch):
    monkeypatch.setattr(premios_routes, "get_concursos_page", lambda **kw: (_stub_concursos(), 1))
    monkeypatch.setattr(premios_routes, "get_empresas", lambda **kw: _stub_empresas())
    monkeypatch.setattr(premios_routes, "_get_sectores", _stub_sectores)
    client = _build_authed_client(monkeypatch)
    resp = client.get("/premios-concursos/")
    assert resp.status_code == 200
    assert b"Premios y concursos" in resp.data
    assert b"SEGURIDAD" in resp.data


def test_premios_resultados_listado_ok(monkeypatch):
    monkeypatch.setattr(
        premios_routes,
        "get_resultados_page",
        lambda **kw: (
            [
                {
                    "id": 7,
                    "periodo_label": "2026-01",
                    "apellido": "Lopez",
                    "nombre": "Ana",
                    "dni": "123",
                    "legajo": "L001",
                    "legajo_snapshot": "L001",
                    "concurso_codigo": "SEGURIDAD",
                    "concurso_nombre": "Premio de seguridad",
                    "concurso_alcance": "global",
                    "concurso_codigo_snapshot": "SEGURIDAD",
                    "concurso_nombre_snapshot": "Premio de seguridad",
                    "sector_nombre": "Logistica",
                    "concurso_sector_nombre": None,
                    "ranking": 1,
                    "observaciones": None,
                }
            ],
            1,
        ),
    )
    monkeypatch.setattr(
        premios_routes,
        "get_resultados_summary",
        lambda **kw: {"total": 1, "primeros": 1, "podios": 1, "empleados": 1, "concursos": 1},
    )
    monkeypatch.setattr(premios_routes, "get_empresas", lambda **kw: _stub_empresas())
    monkeypatch.setattr(premios_routes, "_get_sectores", _stub_sectores)
    monkeypatch.setattr(premios_routes, "get_sucursales", lambda **kw: [])
    monkeypatch.setattr(premios_routes, "get_concursos_for_empresa", lambda *a, **kw: _stub_concursos())
    monkeypatch.setattr(premios_routes, "get_empleados", lambda **kw: _stub_empleados())
    client = _build_authed_client(monkeypatch)
    resp = client.get("/premios-concursos/resultados?empresa_id=1")
    assert resp.status_code == 200
    assert b"Resultados de premios" in resp.data
    assert b"Premio de seguridad" in resp.data


def test_premios_resultados_importar_ok(monkeypatch):
    monkeypatch.setattr(premios_routes, "get_empresas", lambda **kw: _stub_empresas())
    monkeypatch.setattr(
        premios_routes,
        "importar_premios_desde_archivo",
        lambda empresa_id, stream, filename, **kw: {
            "total_filas": 1,
            "importadas": 1,
            "creados": 1,
            "actualizados": 0,
            "errores": 0,
            "detalle_errores": [],
        },
    )
    monkeypatch.setattr(premios_routes, "log_audit", lambda *a, **kw: None)
    client = _build_authed_client(monkeypatch)
    resp = client.post(
        "/premios-concursos/resultados/importar",
        data={
            "empresa_id": "1",
            "archivo": (io.BytesIO(b"periodo,legajo,codigo_concurso,ranking\n"), "premios.csv"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert b"Importacion completada" in resp.data


def test_premios_resultados_export_xlsx_ok(monkeypatch):
    monkeypatch.setattr(
        premios_routes,
        "get_resultados_export",
        lambda **kw: [
            {
                "id": 7,
                "empresa_nombre": "Acme",
                "periodo_label": "2026-01",
                "apellido": "Lopez",
                "nombre": "Ana",
                "dni": "123",
                "legajo": "L001",
                "legajo_snapshot": "L001",
                "concurso_codigo": "SEGURIDAD",
                "concurso_nombre": "Premio de seguridad",
                "concurso_alcance": "global",
                "concurso_codigo_snapshot": "SEGURIDAD",
                "concurso_nombre_snapshot": "Premio de seguridad",
                "sector_nombre": "Logistica",
                "ranking": 1,
                "observaciones": None,
            }
        ],
    )
    monkeypatch.setattr(
        premios_routes,
        "get_resultados_summary",
        lambda **kw: {"total": 1, "primeros": 1, "podios": 1, "empleados": 1, "concursos": 1},
    )
    monkeypatch.setattr(premios_routes, "get_empresas", lambda **kw: _stub_empresas())
    monkeypatch.setattr(premios_routes, "_get_sectores", _stub_sectores)
    monkeypatch.setattr(premios_routes, "get_sucursales", lambda **kw: [])
    monkeypatch.setattr(premios_routes, "get_concursos_for_empresa", lambda *a, **kw: _stub_concursos())
    monkeypatch.setattr(premios_routes, "get_empleados", lambda **kw: _stub_empleados())
    client = _build_authed_client(monkeypatch)

    resp = client.get("/premios-concursos/resultados/export.xlsx?empresa_id=1")

    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp.headers["Content-Type"]
    assert "premios_resultados_" in resp.headers["Content-Disposition"]

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb["Resultados"]
    assert ws["A1"].value == "Resultados de premios"
    assert ws["A16"].value == "Resultados"
    assert ws["B16"].value == 1
    assert ws["A28"].value == "ID"
    assert ws["B29"].value == "Acme"
