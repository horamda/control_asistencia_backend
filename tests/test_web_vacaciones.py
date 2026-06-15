import io

from openpyxl import load_workbook

import app as app_module
import web.auth.decorators as auth_decorators
import web.vacaciones.vacaciones_routes as vacaciones_routes


def _build_client(monkeypatch):
    monkeypatch.setattr(app_module, "init_db", lambda: None)
    app = app_module.create_app()
    app.config["TESTING"] = True
    app.config["WTF_CSRF_ENABLED"] = False
    return app.test_client()


def _login(client, role="admin"):
    with client.session_transaction() as sess:
        sess["user_id"] = 99
        sess["role"] = role
        sess["nombre"] = "Test"


def _build_authed_client(monkeypatch):
    monkeypatch.setattr(auth_decorators, "has_role", lambda user_id, role: True)
    client = _build_client(monkeypatch)
    _login(client)
    return client


def _stub_empleados():
    return [{"id": 10, "nombre": "Ana", "apellido": "Lopez", "dni": "12345", "sector_nombre": "Ventas"}]


def _stub_sectores():
    return [{"id": 4, "nombre": "Ventas", "empresa_nombre": "Acme"}]


def _stub_summary():
    return {
        "total": 1,
        "pendientes": 1,
        "aprobados": 0,
        "rechazados": 0,
        "revertidos": 0,
        "dias_tomados": 0,
        "dias_pendientes": 5,
        "dias_compensatorios": 0,
        "dias_ajustes": 0,
    }


def _stub_saldo():
    return {
        "anio": 2026,
        "empleado": {"id": 10, "dni": "12345", "nombre": "Ana Lopez"},
        "vacaciones": {
            "fecha_ingreso": "2020-01-01",
            "antiguedad_al_31_12": 6,
            "dias_habiles_anio": 261,
            "dias_trabajados_anio": 0,
            "aplica_control_proporcional": False,
            "calculo_proporcional": False,
            "dias_base": 21,
            "dias_compensatorios": 0,
            "dias_ajustes": 0,
            "dias_tomados": 0,
            "dias_pendientes": 5,
            "dias_corresponden": 21,
            "dias_disponibles": 21,
            "dias_disponibles_con_pendientes": 16,
        },
    }


def test_vacaciones_listado_ok(monkeypatch):
    monkeypatch.setattr(
        vacaciones_routes,
        "get_movimientos_page",
        lambda **kw: (
            [
                {
                    "id": 5,
                    "apellido": "Lopez",
                    "nombre": "Ana",
                    "dni": "12345",
                    "empresa_nombre": "Acme",
                    "anio": 2026,
                    "tipo": "tomado",
                    "dias": 5,
                    "estado": "pendiente",
                    "fecha_desde": "2026-01-10",
                    "fecha_hasta": "2026-01-14",
                    "observacion": "Solicitud mobile",
                }
            ],
            1,
        ),
    )
    monkeypatch.setattr(vacaciones_routes, "get_movimientos_summary", lambda **kw: _stub_summary())
    monkeypatch.setattr(vacaciones_routes, "get_empleados", lambda **kw: _stub_empleados())
    monkeypatch.setattr(vacaciones_routes, "get_sectores", lambda **kw: _stub_sectores())
    monkeypatch.setattr(vacaciones_routes, "get_all", lambda: [])
    monkeypatch.setattr(vacaciones_routes, "calcular_resumen_vacaciones", lambda empleado_id, anio: _stub_saldo())

    client = _build_authed_client(monkeypatch)
    resp = client.get("/vacaciones/?empleado_id=10&anio=2026")

    assert resp.status_code == 200
    assert b"Vacaciones" in resp.data
    assert b"Solicitud mobile" in resp.data
    assert b"Aprobar" in resp.data


def test_vacaciones_listado_envia_filtros(monkeypatch):
    captured = {}

    def _fake_get_page(**kw):
        captured.update(kw)
        return ([], 0)

    monkeypatch.setattr(vacaciones_routes, "get_movimientos_page", _fake_get_page)
    monkeypatch.setattr(vacaciones_routes, "get_movimientos_summary", lambda **kw: _stub_summary())
    monkeypatch.setattr(vacaciones_routes, "get_empleados", lambda **kw: _stub_empleados())
    monkeypatch.setattr(vacaciones_routes, "get_sectores", lambda **kw: _stub_sectores())
    monkeypatch.setattr(vacaciones_routes, "get_all", lambda: [])
    monkeypatch.setattr(vacaciones_routes, "calcular_resumen_vacaciones", lambda empleado_id, anio: _stub_saldo())

    client = _build_authed_client(monkeypatch)
    resp = client.get("/vacaciones/?empleado_id=10&sector_id=4&anio=2026&mes=1&tipo=tomado&estado=pendiente&q=ana")

    assert resp.status_code == 200
    assert captured["empleado_id"] == 10
    assert captured["sector_id"] == 4
    assert captured["anio"] == 2026
    assert captured["mes"] == 1
    assert captured["tipo"] == "tomado"
    assert captured["estado"] == "pendiente"
    assert captured["search"] == "ana"


def test_vacaciones_reporte_export_xlsx_ok(monkeypatch):
    client = _build_authed_client(monkeypatch)
    monkeypatch.setattr(auth_decorators, "has_role", lambda user_id, role: True)
    monkeypatch.setattr(
        vacaciones_routes,
        "get_empleados",
        lambda include_inactive=True: [
            {
                "id": 10,
                "nombre": "Ana",
                "apellido": "Lopez",
                "dni": "12345",
                "legajo": "L-10",
                "sector_id": 4,
                "sector_nombre": "Ventas",
                "empresa_nombre": "Acme",
                "puesto_nombre": "Vendedora",
                "activo": 1,
            }
        ],
    )
    monkeypatch.setattr(vacaciones_routes, "get_sectores", lambda include_inactive=True: [{"id": 4, "nombre": "Ventas", "empresa_nombre": "Acme"}])
    monkeypatch.setattr(vacaciones_routes, "calcular_resumen_vacaciones", lambda empleado_id, anio: _stub_saldo())

    resp = client.get("/vacaciones/reporte/export.xlsx?anio=2026&sector_id=4&activo=1&q=ana")

    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp.headers["Content-Type"]
    assert "vacaciones_reporte_2026_" in resp.headers["Content-Disposition"]

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb["Reporte"]
    assert ws["A1"].value == "Reporte de vacaciones"
    assert ws["A5"].value == "Anio"
    assert ws["B6"].value == "Ventas"
    assert ws["A12"].value == "Empleados"
    assert ws["B12"].value == 1
    assert ws["D24"].value == "Lopez Ana"


def test_vacaciones_movimiento_nuevo_post(monkeypatch):
    captured = {}
    monkeypatch.setattr(vacaciones_routes, "get_empleados", lambda **kw: _stub_empleados())
    monkeypatch.setattr(
        vacaciones_routes,
        "crear_movimiento_vacaciones_admin",
        lambda data: captured.update(data) or 77,
    )
    monkeypatch.setattr(vacaciones_routes, "log_audit", lambda *args, **kwargs: None)

    client = _build_authed_client(monkeypatch)
    resp = client.post(
        "/vacaciones/movimientos/nuevo",
        data={
            "empleado_id": "10",
            "anio": "2026",
            "tipo": "compensatorio",
            "dias": "2",
            "estado": "aprobado",
            "observacion": "Temporada",
        },
    )

    assert resp.status_code == 302
    assert captured["empleado_id"] == 10
    assert captured["tipo"] == "compensatorio"
    assert "msg=" in resp.headers["Location"]


def test_vacaciones_aprobar_redirige_con_msg(monkeypatch):
    captured = {}
    monkeypatch.setattr(
        vacaciones_routes,
        "aprobar_movimiento_vacaciones",
        lambda movimiento_id, **kw: captured.update({"movimiento_id": movimiento_id, "actor_id": kw.get("actor_id")}),
    )
    monkeypatch.setattr(vacaciones_routes, "log_audit", lambda *args, **kwargs: None)

    client = _build_authed_client(monkeypatch)
    resp = client.post("/vacaciones/movimientos/aprobar/5")

    assert resp.status_code == 302
    assert captured == {"movimiento_id": 5, "actor_id": 99}
    assert "msg=" in resp.headers["Location"]


def test_vacaciones_rechazar_redirige_con_msg(monkeypatch):
    captured = {}
    monkeypatch.setattr(
        vacaciones_routes,
        "rechazar_movimiento_vacaciones",
        lambda movimiento_id, **kw: captured.update({"movimiento_id": movimiento_id, "actor_id": kw.get("actor_id")}),
    )
    monkeypatch.setattr(vacaciones_routes, "log_audit", lambda *args, **kwargs: None)

    client = _build_authed_client(monkeypatch)
    resp = client.post("/vacaciones/movimientos/rechazar/5")

    assert resp.status_code == 302
    assert captured == {"movimiento_id": 5, "actor_id": 99}
    assert "msg=" in resp.headers["Location"]


def test_vacaciones_cancelar_redirige_con_msg(monkeypatch):
    captured = {}
    monkeypatch.setattr(
        vacaciones_routes,
        "cancelar_movimiento_vacaciones",
        lambda movimiento_id, **kw: captured.update({
            "movimiento_id": movimiento_id,
            "actor_id": kw.get("actor_id"),
            "motivo": kw.get("motivo"),
        }) or 88,
    )
    monkeypatch.setattr(vacaciones_routes, "log_audit", lambda *args, **kwargs: None)

    client = _build_authed_client(monkeypatch)
    resp = client.post("/vacaciones/movimientos/cancelar/5", data={"motivo": "Error de carga"})

    assert resp.status_code == 302
    assert captured == {"movimiento_id": 5, "actor_id": 99, "motivo": "Error de carga"}
    assert "msg=" in resp.headers["Location"]


def test_vacaciones_movimiento_editar_post(monkeypatch):
    captured = {}
    monkeypatch.setattr(
        vacaciones_routes,
        "get_movimiento_by_id",
        lambda movimiento_id: {
            "id": movimiento_id,
            "empleado_id": 10,
            "anio": 2026,
            "tipo": "tomado",
            "dias": 5,
            "estado": "pendiente",
            "fecha_desde": "2026-01-10",
            "fecha_hasta": "2026-01-14",
            "observacion": "Solicitud",
        },
    )
    monkeypatch.setattr(vacaciones_routes, "get_empleados", lambda **kw: _stub_empleados())
    monkeypatch.setattr(
        vacaciones_routes,
        "editar_movimiento_vacaciones_pendiente",
        lambda movimiento_id, data: captured.update({"movimiento_id": movimiento_id, **data}),
    )
    monkeypatch.setattr(vacaciones_routes, "log_audit", lambda *args, **kwargs: None)

    client = _build_authed_client(monkeypatch)
    resp = client.post(
        "/vacaciones/movimientos/editar/5",
        data={
            "empleado_id": "10",
            "anio": "2026",
            "tipo": "tomado",
            "dias": "5",
            "estado": "pendiente",
            "fecha_desde": "2026-01-10",
            "fecha_hasta": "2026-01-14",
            "observacion": "Editada",
        },
    )

    assert resp.status_code == 302
    assert captured["movimiento_id"] == 5
    assert captured["tipo"] == "tomado"
    assert captured["observacion"] == "Editada"


def test_vacaciones_movimientos_export_csv(monkeypatch):
    captured = {}

    def _fake_export(**kw):
        captured.update(kw)
        return [
            {
                "id": 5,
                "empresa_nombre": "Acme",
                "sector_nombre": "Ventas",
                "apellido": "Lopez",
                "nombre": "Ana",
                "dni": "12345",
                "anio": 2026,
                "tipo": "tomado",
                "dias": 5,
                "estado": "aprobado",
                "fecha_desde": "2026-01-10",
                "fecha_hasta": "2026-01-14",
                "observacion": "Descanso",
                "created_at": "2026-01-01",
            }
        ]

    monkeypatch.setattr(vacaciones_routes, "get_movimientos_export", _fake_export)

    client = _build_authed_client(monkeypatch)
    resp = client.get("/vacaciones/movimientos/export.csv?sector_id=4&anio=2026&mes=1")
    text = resp.data.decode("utf-8-sig")

    assert resp.status_code == 200
    assert resp.mimetype == "text/csv"
    assert captured["sector_id"] == 4
    assert captured["anio"] == 2026
    assert captured["mes"] == 1
    assert "id,empresa,sector,empleado" in text
    assert "5,Acme,Ventas,Lopez Ana" in text


def _reporte_empleados():
    return [
        {
            "id": 10,
            "activo": 1,
            "estado": "activo",
            "empresa_nombre": "Acme",
            "sector_id": 4,
            "sector_nombre": "Ventas",
            "puesto_nombre": "Gerente",
            "apellido": "Lopez",
            "nombre": "Ana",
            "dni": "12345",
            "legajo": "L10",
        },
        {
            "id": 11,
            "activo": 1,
            "estado": "activo",
            "empresa_nombre": "Acme",
            "sector_id": 5,
            "sector_nombre": "Operaciones",
            "puesto_nombre": "Chofer",
            "apellido": "Gomez",
            "nombre": "Luis",
            "dni": "22222",
            "legajo": "L11",
        },
        {
            "id": 12,
            "activo": 0,
            "estado": "inactivo",
            "empresa_nombre": "Acme",
            "sector_id": 4,
            "sector_nombre": "Ventas",
            "puesto_nombre": "Analista",
            "apellido": "Perez",
            "nombre": "Marta",
            "dni": "33333",
            "legajo": "L12",
        },
    ]


def _reporte_resumen(empleado_id, anio):
    return {
        "anio": anio,
        "empleado": {"id": empleado_id, "dni": "12345", "nombre": "Ana Lopez"},
        "vacaciones": {
            "fecha_ingreso": "2020-01-01",
            "antiguedad_al_31_12": 6,
            "dias_habiles_anio": 261,
            "dias_trabajados_anio": 180,
            "aplica_control_proporcional": False,
            "calculo_proporcional": False,
            "dias_base": 21,
            "dias_compensatorios": 2,
            "dias_ajustes": 0,
            "dias_tomados": 5,
            "dias_pendientes": 3,
            "dias_corresponden": 23,
            "dias_disponibles": 18,
            "dias_disponibles_con_pendientes": 15,
        },
    }


def test_vacaciones_reporte_filtra_por_area_y_activos(monkeypatch):
    captured = []
    monkeypatch.setattr(vacaciones_routes, "get_empleados", lambda **kw: _reporte_empleados())
    monkeypatch.setattr(vacaciones_routes, "get_sectores", lambda **kw: _stub_sectores())

    def _fake_resumen(empleado_id, anio):
        captured.append((empleado_id, anio))
        return _reporte_resumen(empleado_id, anio)

    monkeypatch.setattr(vacaciones_routes, "calcular_resumen_vacaciones", _fake_resumen)

    client = _build_authed_client(monkeypatch)
    resp = client.get("/vacaciones/reporte?sector_id=4&activo=1&anio=2026")

    assert resp.status_code == 200
    assert captured == [(10, 2026)]
    assert b"Reporte de vacaciones" in resp.data
    assert b"Lopez Ana" in resp.data
    assert b"Gomez Luis" not in resp.data
    assert b"Perez Marta" not in resp.data
    assert b"Compensatorios" in resp.data
    assert b"Pendientes por tomar" in resp.data
    assert b"18.00" in resp.data


def test_vacaciones_reporte_export_csv(monkeypatch):
    monkeypatch.setattr(vacaciones_routes, "get_empleados", lambda **kw: _reporte_empleados())
    monkeypatch.setattr(vacaciones_routes, "calcular_resumen_vacaciones", _reporte_resumen)

    client = _build_authed_client(monkeypatch)
    resp = client.get("/vacaciones/reporte/export.csv?sector_id=4&activo=1&anio=2026")
    text = resp.data.decode("utf-8-sig")

    assert resp.status_code == 200
    assert resp.mimetype == "text/csv"
    assert "vacaciones_reporte_2026" in resp.headers["Content-Disposition"]
    assert "empresa,area,puesto,empleado" in text
    assert "dias_corresponden,dias_compensatorios,dias_tomados,dias_pendientes_por_tomar" in text
    assert "Acme,Ventas,Gerente,Lopez Ana" in text
    assert "Total empleados,1" in text
    assert "Dias compensatorios,2.0" in text
    assert "Dias pendientes por tomar,18.0" in text
