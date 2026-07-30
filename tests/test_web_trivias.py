import io
import datetime

from openpyxl import load_workbook

import app as app_module
import web.auth.decorators as auth_decorators
import web.trivias.trivia_admin_routes as trivia_admin_routes


_TRIVIA = {
    "id": 3,
    "titulo": "Trivia Mayo 2026",
    "descripcion": "Preguntas operativas",
    "fecha_inicio": datetime.datetime(2026, 5, 24, 8, 0),
    "fecha_fin": datetime.datetime(2026, 5, 31, 23, 59),
    "estado": "activa",
    "premio": "Vale",
    "mensaje_ganador": None,
    "anio": 2026,
}


def _build_client(monkeypatch):
    monkeypatch.setattr(app_module, "init_db", lambda: None)
    app = app_module.create_app()
    app.config["TESTING"] = True
    app.config["WTF_CSRF_ENABLED"] = False
    return app.test_client()


def _login(client):
    with client.session_transaction() as sess:
        sess["user_id"] = 99
        sess["role"] = "admin"
        sess["nombre"] = "Test"


def _build_authed_client(monkeypatch):
    monkeypatch.setattr(auth_decorators, "has_role", lambda user_id, role: True)
    client = _build_client(monkeypatch)
    _login(client)
    return client


def test_trivia_resultados_admin_ok(monkeypatch):
    monkeypatch.setattr(trivia_admin_routes.repo, "get_trivia_by_id", lambda tid: _TRIVIA)
    monkeypatch.setattr(trivia_admin_routes, "get_sucursales", lambda **kw: [])
    monkeypatch.setattr(
        trivia_admin_routes.repo,
        "get_resultados_admin_trivia",
        lambda tid, sucursal_id=None: [
            {
                "empleado_id": 10,
                "empleado_dni": "30111222",
                "empleado_legajo": "L001",
                "empleado_nombre": "Ana",
                "empleado_apellido": "Lopez",
                "empleado_nombre_completo": "Lopez Ana",
                "empleado_activo": 1,
                "sector_nombre": "Logistica",
                "habilitado_por_alcance": 1,
                "resultado_id": 1,
                "fecha_inicio_participacion": datetime.datetime(2026, 5, 25, 9, 0),
                "fecha_finalizacion": datetime.datetime(2026, 5, 25, 9, 2),
                "tiempo_total_segundos": 120,
                "puntos_total": 25,
                "correctas": 2,
                "incorrectas": 0,
                "posicion": None,
                "es_ganador": 0,
                "estado_resultado": "completado",
                "exclusion_id": None,
                "exclusion_motivo": None,
                "exclusion_creado_en": None,
            },
            {
                "empleado_id": 11,
                "empleado_dni": "30222333",
                "empleado_legajo": "L002",
                "empleado_nombre": "Luis",
                "empleado_apellido": "Gomez",
                "empleado_nombre_completo": "Gomez Luis",
                "empleado_activo": 1,
                "sector_nombre": "Logistica",
                "habilitado_por_alcance": 1,
                "resultado_id": None,
                "fecha_inicio_participacion": None,
                "fecha_finalizacion": None,
                "tiempo_total_segundos": None,
                "puntos_total": None,
                "correctas": None,
                "incorrectas": None,
                "posicion": None,
                "es_ganador": 0,
                "estado_resultado": None,
                "exclusion_id": 4,
                "exclusion_motivo": "No participa",
                "exclusion_creado_en": datetime.datetime(2026, 5, 26, 10, 0),
            },
        ],
    )
    monkeypatch.setattr(
        trivia_admin_routes.repo,
        "get_exclusiones_trivia",
        lambda tid: [
            {
                "empleado_id": 11,
                "empleado_dni": "30222333",
                "empleado_nombre_completo": "Gomez Luis",
                "sector_nombre": "Logistica",
                "motivo": "No participa",
            }
        ],
    )
    monkeypatch.setattr(trivia_admin_routes.repo, "get_respuestas_admin_trivia", lambda tid: [])
    monkeypatch.setattr(
        trivia_admin_routes,
        "get_empleados",
        lambda include_inactive=False: [{"id": 10, "apellido": "Lopez", "nombre": "Ana", "dni": "30111222"}],
    )
    monkeypatch.setattr(
        trivia_admin_routes,
        "calcular_ranking",
        lambda tid: [
            {
                "empleado_id": 10,
                "posicion": 1,
                "es_ganador": True,
            }
        ],
    )
    client = _build_authed_client(monkeypatch)

    resp = client.get("/admin/trivias/3/resultados")

    assert resp.status_code == 200
    assert b"Resultados" in resp.data
    assert b"Lopez Ana" in resp.data
    assert b"Gomez Luis" in resp.data
    assert b"Excluido" in resp.data


def test_trivia_agregar_exclusion_recalcula(monkeypatch):
    llamadas = []
    recalculos = []
    monkeypatch.setattr(trivia_admin_routes.repo, "get_trivia_by_id", lambda tid: _TRIVIA)
    monkeypatch.setattr(
        trivia_admin_routes.repo,
        "add_exclusion_trivia",
        lambda tid, eid, motivo=None, creado_por=None: llamadas.append((tid, eid, motivo, creado_por)),
    )
    monkeypatch.setattr(
        trivia_admin_routes,
        "recalcular_resultados_trivia",
        lambda tid: recalculos.append(tid),
    )
    monkeypatch.setattr(trivia_admin_routes, "log_audit", lambda *a, **kw: None)
    client = _build_authed_client(monkeypatch)

    resp = client.post(
        "/admin/trivias/3/exclusiones",
        data={"empleado_id": "11", "motivo": "No participa"},
    )

    assert resp.status_code == 302
    assert llamadas == [(3, 11, "No participa", 99)]
    assert recalculos == [3]


def test_trivia_ranking_anual_muestra_exclusiones(monkeypatch):
    monkeypatch.setattr(
        trivia_admin_routes,
        "calcular_ranking_anual",
        lambda anio: [
            {
                "empleado_id": 10,
                "empleado_dni": "30111222",
                "empleado_nombre": "Lopez Ana",
                "puntos_anuales": 80,
                "trivias_participadas": 2,
                "trivias_ganadas": 1,
                "correctas_totales": 8,
                "incorrectas_totales": 2,
                "tiempo_total_anual": 120,
                "posicion": 1,
                "es_ganador_anual": 1,
            }
        ],
    )
    monkeypatch.setattr(
        trivia_admin_routes.repo,
        "get_exclusiones_ranking_anual",
        lambda anio: [
            {
                "empleado_id": 11,
                "empleado_dni": "30222333",
                "empleado_nombre_completo": "Gomez Luis",
                "sector_nombre": "Logistica",
                "motivo": "No computa anual",
            }
        ],
    )
    monkeypatch.setattr(
        trivia_admin_routes,
        "get_empleados",
        lambda include_inactive=False: [{"id": 10, "apellido": "Lopez", "nombre": "Ana", "dni": "30111222"}],
    )
    client = _build_authed_client(monkeypatch)

    resp = client.get("/admin/trivias/ranking-anual?anio=2026")

    assert resp.status_code == 200
    assert b"Ranking anual 2026" in resp.data
    assert b"Exclusiones del ranking anual" in resp.data
    assert b"Gomez Luis" in resp.data
    assert b"No computa anual" in resp.data


def test_trivia_agregar_exclusion_anual_recalcula(monkeypatch):
    llamadas = []
    recalculos = []
    monkeypatch.setattr(
        trivia_admin_routes.repo,
        "add_exclusion_ranking_anual",
        lambda anio, eid, motivo=None, creado_por=None: llamadas.append((anio, eid, motivo, creado_por)),
    )
    monkeypatch.setattr(
        trivia_admin_routes.repo,
        "recalcular_ranking_anual",
        lambda anio: recalculos.append(anio),
    )
    monkeypatch.setattr(trivia_admin_routes, "log_audit", lambda *a, **kw: None)
    client = _build_authed_client(monkeypatch)

    resp = client.post(
        "/admin/trivias/ranking-anual/exclusiones",
        data={"anio": "2026", "empleado_id": "11", "motivo": "No computa anual"},
    )

    assert resp.status_code == 302
    assert llamadas == [(2026, 11, "No computa anual", 99)]
    assert recalculos == [2026]


def test_trivia_resultados_export_xlsx_ok(monkeypatch):
    monkeypatch.setattr(trivia_admin_routes.repo, "get_trivia_by_id", lambda tid: _TRIVIA)
    monkeypatch.setattr(
        trivia_admin_routes.repo,
        "get_resultados_admin_trivia",
        lambda tid, sucursal_id=None: [
            {
                "empleado_id": 10,
                "empleado_dni": "30111222",
                "empleado_legajo": "L001",
                "empleado_nombre": "Ana",
                "empleado_apellido": "Lopez",
                "empleado_nombre_completo": "Lopez Ana",
                "empleado_activo": 1,
                "sector_nombre": "Logistica",
                "habilitado_por_alcance": 1,
                "resultado_id": 1,
                "fecha_inicio_participacion": datetime.datetime(2026, 5, 25, 9, 0),
                "fecha_finalizacion": datetime.datetime(2026, 5, 25, 9, 2),
                "tiempo_total_segundos": 120,
                "puntos_total": 25,
                "correctas": 2,
                "incorrectas": 0,
                "posicion": None,
                "es_ganador": 0,
                "estado_resultado": "completado",
                "exclusion_id": None,
                "exclusion_motivo": None,
                "exclusion_creado_en": None,
            },
            {
                "empleado_id": 11,
                "empleado_dni": "30222333",
                "empleado_legajo": "L002",
                "empleado_nombre": "Luis",
                "empleado_apellido": "Gomez",
                "empleado_nombre_completo": "Gomez Luis",
                "empleado_activo": 1,
                "sector_nombre": "Logistica",
                "habilitado_por_alcance": 1,
                "resultado_id": None,
                "fecha_inicio_participacion": None,
                "fecha_finalizacion": None,
                "tiempo_total_segundos": None,
                "puntos_total": None,
                "correctas": None,
                "incorrectas": None,
                "posicion": None,
                "es_ganador": 0,
                "estado_resultado": None,
                "exclusion_id": 4,
                "exclusion_motivo": "No participa",
                "exclusion_creado_en": datetime.datetime(2026, 5, 26, 10, 0),
            },
        ],
    )
    monkeypatch.setattr(trivia_admin_routes, "calcular_ranking", lambda tid: [{"empleado_id": 10, "posicion": 1, "es_ganador": True}])
    client = _build_authed_client(monkeypatch)

    resp = client.get("/admin/trivias/3/resultados/export.xlsx")

    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp.headers["Content-Type"]
    assert "trivia_3_resultados.xlsx" in resp.headers["Content-Disposition"]

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb["Resultados"]
    assert ws["A1"].value == "Resultados de trivia"
    assert ws["A11"].value == "Habilitados"
    assert ws["B11"].value == 1
    assert ws["A20"].value == 3
    assert ws["K20"].value == "completado"
    assert ws["K21"].value == "excluido"


def test_trivia_eliminar_exclusion_anual_recalcula(monkeypatch):
    llamadas = []
    recalculos = []
    monkeypatch.setattr(
        trivia_admin_routes.repo,
        "remove_exclusion_ranking_anual",
        lambda anio, eid: llamadas.append((anio, eid)),
    )
    monkeypatch.setattr(
        trivia_admin_routes.repo,
        "recalcular_ranking_anual",
        lambda anio: recalculos.append(anio),
    )
    monkeypatch.setattr(trivia_admin_routes, "log_audit", lambda *a, **kw: None)
    client = _build_authed_client(monkeypatch)

    resp = client.post(
        "/admin/trivias/ranking-anual/exclusiones/11/eliminar",
        data={"anio": "2026"},
    )

    assert resp.status_code == 302
    assert llamadas == [(2026, 11)]
    assert recalculos == [2026]
