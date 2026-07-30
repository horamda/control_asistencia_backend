import datetime
import io

from openpyxl import load_workbook

import app as app_module
import web.auth.decorators as auth_decorators
import web.kpis_sectoriales.kpis_sectoriales_routes as kpi_routes


def _build_client(monkeypatch):
    monkeypatch.setattr(app_module, "init_db", lambda: None)
    app = app_module.create_app()
    app.config["TESTING"] = True
    app.config["WTF_CSRF_ENABLED"] = False
    return app.test_client()


def _login_session(client):
    with client.session_transaction() as sess:
        sess["user_id"] = 21
        sess["user_role"] = "admin"


def test_kpis_resultados_empleado_renderiza_calendario(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)

    monkeypatch.setattr(kpi_routes, "get_empresas", lambda: [{"id": 1, "razon_social": "Empresa A"}])
    monkeypatch.setattr(kpi_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(
        kpi_routes,
        "get_sectores_page",
        lambda page, per_page, empresa_id=None, activo=None: ([{"id": 2, "nombre": "Ventas"}], 1),
    )
    monkeypatch.setattr(
        kpi_routes,
        "get_empleados_by_sector_para_kpis",
        lambda empresa_id, sector_id, sucursal_id=None: [
            {
                "id": 7,
                "empresa_id": 1,
                "sector_id": 2,
                "legajo": "L-7",
                "dni": "30111222",
                "nombre": "Ana",
                "apellido": "Perez",
                "empresa_nombre": "Empresa A",
                "sector_nombre": "Ventas",
            }
        ],
    )
    monkeypatch.setattr(
        kpi_routes,
        "get_kpis_by_sector",
        lambda sector_id, activo=None: [
            {
                "id": 5,
                "codigo": "VENTAS",
                "nombre": "Ventas cerradas",
                "unidad": "ventas",
                "tipo_acumulacion": "suma",
                "mayor_es_mejor": 1,
                "activo": 1,
            }
        ],
    )
    monkeypatch.setattr(
        kpi_routes,
        "get_resultados_empleado_kpis_anio",
        lambda empleado_id, sector_id, anio: [
            {
                "kpi_id": 5,
                "codigo": "VENTAS",
                "nombre": "Ventas cerradas",
                "unidad": "ventas",
                "tipo_acumulacion": "suma",
                "mayor_es_mejor": 1,
                "objetivo_valor": 3650,
                "condicion": "gte",
                "valor_min": None,
                "valor_max": None,
                "fecha": datetime.date(2026, 5, 1),
                "valor": 20,
            }
        ],
    )

    resp = client.get("/kpis-sectoriales/resultados?empresa_id=1&sector_id=2&empleado_id=7&anio=2026&mes=5")

    assert resp.status_code == 200
    assert b"KPIs diarios y mensuales" in resp.data
    assert b"Calendario de resultados" in resp.data
    assert b"Perez Ana" in resp.data
    assert b"kpi-calendar-name" in resp.data
    assert b"Ventas cerradas" in resp.data
    assert b"VENTAS" in resp.data
    assert b"20" in resp.data
    assert b"Eliminar resultados del mes" in resp.data
    assert b"2026-05" in resp.data


def test_kpis_resultados_eliminar_mes_respeta_filtros(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)

    called = {}

    def fake_delete(empresa_id, sector_id, anio, mes, empleado_id=None, kpi_id=None):
        called.update({
            "empresa_id": empresa_id,
            "sector_id": sector_id,
            "anio": anio,
            "mes": mes,
            "empleado_id": empleado_id,
            "kpi_id": kpi_id,
        })
        return 4

    monkeypatch.setattr(kpi_routes, "delete_resultados_mes", fake_delete)
    monkeypatch.setattr(kpi_routes, "log_audit", lambda *args, **kwargs: None)

    resp = client.post(
        "/kpis-sectoriales/resultados/eliminar-mes",
        data={
            "empresa_id": "1",
            "sector_id": "2",
            "empleado_id": "7",
            "kpi_id": "5",
            "anio": "2026",
            "mes": "5",
            "confirmacion": "2026-05",
        },
        follow_redirects=False,
    )

    assert resp.status_code == 302
    assert called == {
        "empresa_id": 1,
        "sector_id": 2,
        "anio": 2026,
        "mes": 5,
        "empleado_id": 7,
        "kpi_id": 5,
    }


def test_kpis_resultados_eliminar_mes_exige_confirmacion(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)

    def fail_delete(*args, **kwargs):
        raise AssertionError("No debe borrar sin confirmacion exacta")

    monkeypatch.setattr(kpi_routes, "delete_resultados_mes", fail_delete)

    resp = client.post(
        "/kpis-sectoriales/resultados/eliminar-mes",
        data={
            "empresa_id": "1",
            "sector_id": "2",
            "anio": "2026",
            "mes": "5",
            "confirmacion": "mayo",
        },
        follow_redirects=False,
    )

    assert resp.status_code == 302
    assert "error=" in resp.headers["Location"]


def test_kpis_resultados_export_xlsx_ok(monkeypatch):
    client = _build_client(monkeypatch)
    _login_session(client)
    monkeypatch.setattr(auth_decorators, "has_role", lambda actor_id, role: True)

    monkeypatch.setattr(kpi_routes, "get_empresas", lambda: [{"id": 1, "razon_social": "Empresa A"}])
    monkeypatch.setattr(kpi_routes, "get_sucursales", lambda include_inactive=True: [])
    monkeypatch.setattr(
        kpi_routes,
        "get_sectores_page",
        lambda page, per_page, empresa_id=None, activo=None: ([{"id": 2, "nombre": "Ventas"}], 1),
    )
    monkeypatch.setattr(
        kpi_routes,
        "get_empleados_by_sector_para_kpis",
        lambda empresa_id, sector_id, sucursal_id=None: [
            {
                "id": 7,
                "empresa_id": 1,
                "sector_id": 2,
                "legajo": "L-7",
                "dni": "30111222",
                "nombre": "Ana",
                "apellido": "Perez",
                "empresa_nombre": "Empresa A",
                "sector_nombre": "Ventas",
            }
        ],
    )
    monkeypatch.setattr(
        kpi_routes,
        "get_kpis_by_sector",
        lambda sector_id, activo=None: [
            {
                "id": 5,
                "codigo": "VENTAS",
                "nombre": "Ventas cerradas",
                "unidad": "ventas",
                "tipo_acumulacion": "suma",
                "mayor_es_mejor": 1,
                "activo": 1,
            }
        ],
    )
    monkeypatch.setattr(
        kpi_routes,
        "get_resultados_empleado_kpis_anio",
        lambda empleado_id, sector_id, anio: [
            {
                "kpi_id": 5,
                "codigo": "VENTAS",
                "nombre": "Ventas cerradas",
                "unidad": "ventas",
                "tipo_acumulacion": "suma",
                "mayor_es_mejor": 1,
                "objetivo_valor": 3650,
                "condicion": "gte",
                "valor_min": None,
                "valor_max": None,
                "fecha": datetime.date(2026, 5, 1),
                "valor": 20,
            }
        ],
    )

    resp = client.get("/kpis-sectoriales/resultados/export.xlsx?empresa_id=1&sector_id=2&empleado_id=7&anio=2026&mes=5")

    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp.headers["Content-Type"]
    assert "kpis_resultados_2026_05_" in resp.headers["Content-Disposition"]

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb["Resumen"]
    assert ws["A1"].value == "Resultados KPI por empleado"
    assert ws["A15"].value == "KPIs evaluados"
    assert ws["B15"].value == 1
    assert ws["A23"].value == "VENTAS"
    assert ws["B23"].value == "Ventas cerradas"
    assert ws["A27"].value == "2026-05-01"
